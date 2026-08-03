from pathlib import Path

from openpyxl import Workbook, load_workbook

from app.auth import hash_password, verify_password
from app.models import User
from app.services.exports import export_customer_company_workbook, export_total_workbook
from app.services.orders import create_order_line, get_order_balances, preview_excel_orders
from app.services.shipments import submit_shipment_report


def test_password_hash_verifies_plain_password():
    password_hash = hash_password("123456")
    assert password_hash != "123456"
    assert verify_password("123456", password_hash) is True
    assert verify_password("bad", password_hash) is False


def test_worker_shipment_waits_for_review_and_does_not_count_as_shipped(db_session):
    user = User(username="worker_a", display_name="仓库01", password_hash="x", role="worker", is_active=True)
    db_session.add(user)
    db_session.commit()
    create_order_line(db_session, "福建", "小偷", "小偷女款", "M", 300, "2026-07-14", "", "")

    report = submit_shipment_report(db_session, user.id, "2026-07-17", "福建", "小偷", "小偷女款", [{"size": "M", "quantity": 120}], [], "")

    assert report.status == "pending_review"
    assert "等待老板审核" in report.review_reason
    balance = get_order_balances(db_session, "福建")[0]
    assert balance["shipped"] == 0
    assert balance["remaining"] == 300


def test_over_shipment_goes_pending(db_session):
    user = User(username="worker_a", display_name="仓库01", password_hash="x", role="worker", is_active=True)
    db_session.add(user)
    db_session.commit()
    create_order_line(db_session, "福建", "小偷", "小偷女款", "M", 50, "2026-07-14", "", "")

    report = submit_shipment_report(db_session, user.id, "2026-07-17", "福建", "小偷", "小偷女款", [{"size": "M", "quantity": 80}], [], "")

    assert report.status == "pending_review"
    assert "超发" in report.review_reason
    assert get_order_balances(db_session, "福建")[0]["shipped"] == 0


def test_same_style_duplicate_orders_stay_separate_and_shipments_allocate(db_session):
    from app.models import ShipmentLine, ShipmentReport, User

    admin = User(username="batch_admin", display_name="老板", password_hash="x", role="admin", is_active=True)
    db_session.add(admin)
    db_session.commit()
    create_order_line(db_session, "艾润特", "裁判", "1688裁判COS", "L", 1000)
    create_order_line(db_session, "艾润特", "裁判", "1688裁判COS", "L", 2000)
    report = ShipmentReport(
        user_id=admin.id,
        ship_date="2026-07-10",
        company_name="艾润特",
        product_name="裁判",
        style_name="1688裁判COS",
        status="auto_approved",
    )
    db_session.add(report)
    db_session.flush()
    db_session.add(ShipmentLine(report_id=report.id, size="L", quantity=1500))
    db_session.commit()

    rows = get_order_balances(db_session, "艾润特")

    assert [row["ordered"] for row in rows] == [1000, 2000]
    assert [row["shipped"] for row in rows] == [1000, 500]
    assert [row["remaining"] for row in rows] == [0, 1500]
    assert [row["order_ref"] for row in rows] == ["第1单", "第2单"]


def test_assigned_shipment_line_counts_against_its_order_only(db_session):
    from app.models import ShipmentLine, ShipmentReport, User

    admin = User(username="assigned_history_admin", display_name="老板", password_hash="x", role="admin", is_active=True)
    db_session.add(admin)
    db_session.commit()
    first = create_order_line(db_session, "艾润特", "裁判", "圆领裁判", "S", 100, "2026-06-02")
    second = create_order_line(db_session, "艾润特", "裁判", "圆领裁判", "S", 1000, "2026-06-17")
    report = ShipmentReport(
        user_id=admin.id,
        ship_date="历史导入",
        company_name="艾润特",
        product_name="裁判",
        style_name="圆领裁判",
        status="auto_approved",
        note="06-29 发 802件；07-01 发 198件",
    )
    db_session.add(report)
    db_session.flush()
    db_session.add(ShipmentLine(report_id=report.id, order_line_id=second.id, size="S", quantity=1000))
    db_session.commit()

    rows = get_order_balances(db_session, "艾润特")

    assert [(row["order_date"], row["ordered"], row["shipped"], row["remaining"]) for row in rows] == [
        ("2026-06-02", 100, 0, 100),
        ("2026-06-17", 1000, 1000, 0),
    ]


def test_export_keeps_unassigned_shipments_off_orders_filled_by_assigned_history(db_session, tmp_path):
    from app.models import ShipmentLine, ShipmentReport, User

    admin = User(username="history_priority_admin", display_name="老板", password_hash="x", role="admin", is_active=True)
    db_session.add(admin)
    db_session.commit()
    first = create_order_line(db_session, "艾润特", "裁判", "圆领裁判", "L", 1100, "2026-06-02")
    create_order_line(db_session, "艾润特", "裁判", "圆领裁判", "L", 2000, "2026-06-17")
    history = ShipmentReport(
        user_id=admin.id,
        ship_date="历史导入",
        company_name="1688",
        product_name="裁判",
        style_name="1688裁判COS",
        status="auto_approved",
        note="06-23 发 1075件；07-14 发 25件",
    )
    db_session.add(history)
    db_session.flush()
    db_session.add(ShipmentLine(report_id=history.id, order_line_id=first.id, size="L", quantity=1100))
    current = ShipmentReport(
        user_id=admin.id,
        ship_date="2026-07-20",
        company_name="艾润特",
        product_name="裁判",
        style_name="圆领裁判",
        status="auto_approved",
    )
    db_session.add(current)
    db_session.flush()
    db_session.add(ShipmentLine(report_id=current.id, size="L", quantity=166))
    db_session.commit()
    output = tmp_path / "艾润特.xlsx"

    export_customer_company_workbook(db_session, "艾润特", output)

    ws = load_workbook(output)["客户发货明细"]
    assert ws["D2"].value == "2026-06-02"
    assert ws["G2"].value == "06-23 发 1075件；07-14 发 25件"
    assert ws["I2"].value == 1100
    assert ws["D3"].value == "2026-06-17"
    assert ws["G3"].value == "2026-07-20 发 166件"
    assert ws["I3"].value == 166


def test_preview_excel_orders_reads_current_total_format(tmp_path: Path):
    file_path = tmp_path / "orders.xlsx"
    wb = Workbook()
    ws = wb.active
    ws.title = "订单发货明细"
    ws.append(["标题"])
    ws.append([])
    ws.append([])
    ws.append(["公司", "产品", "款式", "尺码", "发货记录", "本次发货", "已发合计", "未发数量", "下单数量", "状态", "备注", "订单批次"])
    ws.append(["福建", "小偷", "小偷女款", "M", "", "", 0, 300, 300, "未发", "测试", "批次1"])
    wb.save(file_path)

    preview = preview_excel_orders(file_path)

    assert preview["valid_rows"][0]["company"] == "福建"
    assert preview["valid_rows"][0]["quantity"] == 300
    assert preview["errors"] == []


def test_total_export_contains_balance_columns(db_session, tmp_path):
    create_order_line(db_session, "福建", "小偷", "小偷女款", "M", 300, "2026-07-14", "", "备注")
    output = tmp_path / "总表.xlsx"

    export_total_workbook(db_session, output)

    wb = load_workbook(output)
    ws = wb["订单发货明细"]
    headers = [cell.value for cell in ws[1]]
    assert headers[:11] == ["公司", "产品", "款式", "订单", "尺码", "SKU", "发货明细", "下单数量", "已发数量", "未发数量", "超发数量"]
    assert ws["A2"].value == "福建"
    assert ws["J2"].value == 300


def test_export_balance_sheet_shows_shipment_detail_not_note(db_session, tmp_path):
    create_order_line(db_session, "源兴发", "小红帽", "小红帽男", "M", 300, "2026-07-14", "", "内部备注")
    output = tmp_path / "源兴发表.xlsx"

    export_total_workbook(db_session, output)

    wb = load_workbook(output)
    ws = wb["订单发货明细"]
    headers = [cell.value for cell in ws[1]]
    assert "发货明细" in headers
    assert "备注" not in headers


def test_customer_company_export_hides_internal_review_fields(db_session, tmp_path):
    admin = User(username="boss", display_name="老板", password_hash="x", role="admin", is_active=True)
    db_session.add(admin)
    db_session.commit()
    create_order_line(db_session, "源兴发", "小红帽", "小红帽男", "M", 300, "2026-07-14", "", "内部备注")
    submit_shipment_report(
        db_session,
        admin.id,
        "2026-07-15",
        "源兴发",
        "小红帽",
        "小红帽男",
        [{"size": "M", "quantity": 80}],
        [],
        "仓库内部说明",
    )
    output = tmp_path / "源兴发客户版.xlsx"

    export_customer_company_workbook(db_session, "源兴发", output)

    wb = load_workbook(output)
    assert wb.sheetnames == ["客户发货明细"]
    ws = wb["客户发货明细"]
    headers = [cell.value for cell in ws[1]]
    assert headers[:10] == ["公司", "产品", "款式", "订单", "尺码", "SKU", "发货明细", "下单数量", "已发数量", "未发数量"]
    visible_headers = [header for header in headers if header]
    assert "上报人" not in visible_headers
    assert "状态" not in visible_headers
    assert "异常原因" not in visible_headers
    assert "备注" not in visible_headers
    assert ws["G2"].value == "2026-07-15 发 80件"


def test_customer_export_splits_shipment_details_across_duplicate_orders(db_session, tmp_path):
    admin = User(username="split_export_boss", display_name="老板", password_hash="x", role="admin", is_active=True)
    db_session.add(admin)
    db_session.commit()
    create_order_line(db_session, "艾润特", "裁判", "圆领裁判", "L", 100, "2026-06-17")
    create_order_line(db_session, "艾润特", "裁判", "圆领裁判", "L", 100, "2026-06-23")
    submit_shipment_report(
        db_session,
        admin.id,
        "2026-07-20",
        "艾润特",
        "裁判",
        "圆领裁判",
        [{"size": "L", "quantity": 130}],
        [],
        "",
    )
    submit_shipment_report(
        db_session,
        admin.id,
        "2026-07-21",
        "艾润特",
        "裁判",
        "圆领裁判",
        [{"size": "L", "quantity": 20}],
        [],
        "",
    )
    output = tmp_path / "艾润特客户版.xlsx"

    export_customer_company_workbook(db_session, "艾润特", output)

    ws = load_workbook(output)["客户发货明细"]
    split_text = "2026-07-20 发 130件，其中100件发2026-06-17订单，30件发2026-06-23订单"
    assert ws["D2"].value == "2026-06-17"
    assert ws["G2"].value == split_text
    assert ws["I2"].value == 100
    assert ws["D3"].value == "2026-06-23"
    assert ws["G3"].value == f"{split_text}；2026-07-21 发 20件"
    assert ws["I3"].value == 50


def test_customer_export_shipment_detail_uses_canonical_names(db_session, tmp_path):
    from app.models import ProductAlias

    admin = User(username="canonical_boss", display_name="老板", password_hash="x", role="admin", is_active=True)
    db_session.add(admin)
    db_session.add(
        ProductAlias(
            company_name="合肥Hoo",
            alias_product="聪明的红帽子",
            alias_style="小红帽",
            canonical_product="小红帽",
            canonical_style="小红帽男款",
        )
    )
    db_session.commit()
    create_order_line(db_session, "合肥Hoo", "聪明的红帽子", "小红帽", "M", 400)
    submit_shipment_report(
        db_session,
        admin.id,
        "2026-07-19",
        "合肥Hoo",
        "小红帽",
        "小红帽男款",
        [{"size": "M", "quantity": 317}],
        [],
        "",
    )
    output = tmp_path / "合肥客户版.xlsx"

    export_customer_company_workbook(db_session, "合肥Hoo", output)

    ws = load_workbook(output)["客户发货明细"]
    assert ws["G2"].value == "2026-07-19 发 317件"


def test_customer_export_shipment_detail_canonicalizes_existing_reports(db_session, tmp_path):
    from app.models import ProductAlias, ShipmentLine, ShipmentReport

    admin = User(username="legacy_alias_boss", display_name="老板", password_hash="x", role="admin", is_active=True)
    db_session.add(admin)
    db_session.add(
        ProductAlias(
            company_name="张鹏",
            alias_product="女士赛车服",
            alias_style="女士赛车服",
            canonical_product="赛车服",
            canonical_style="女士赛车服",
        )
    )
    db_session.commit()
    create_order_line(db_session, "张鹏", "赛车服", "女士赛车服", "L", 400)
    report = ShipmentReport(
        user_id=admin.id,
        ship_date="2026-07-17",
        company_name="张鹏",
        product_name="女士赛车服",
        style_name="女士赛车服",
        status="approved_after_edit",
    )
    db_session.add(report)
    db_session.flush()
    db_session.add(ShipmentLine(report_id=report.id, size="L", quantity=180))
    db_session.commit()
    output = tmp_path / "张鹏客户版.xlsx"

    export_customer_company_workbook(db_session, "张鹏", output)

    ws = load_workbook(output)["客户发货明细"]
    assert ws["G2"].value == "2026-07-17 发 180件"

from pathlib import Path

from openpyxl import Workbook, load_workbook

from app.services.exports import export_company_workbook
from app.services.orders import create_order_line, get_order_balances, update_order_line
from app.services.skus import import_sku_mappings_from_excel, upsert_sku_mapping


def test_import_sku_mappings_from_excel_and_show_in_balances(db_session, tmp_path: Path):
    create_order_line(db_session, "源兴发", "僵尸棒球", "僵尸棒球", "L", 400)
    file_path = tmp_path / "源兴发SKU对应表.xlsx"
    wb = Workbook()
    ws = wb.active
    ws.title = "源兴发SKU对应表"
    ws.append(["公司", "产品", "款式", "尺码", "SKU"])
    ws.append(["源兴发", "僵尸棒球", "僵尸棒球", "L", "SKU-ZOMBIE-L"])
    wb.save(file_path)

    result = import_sku_mappings_from_excel(db_session, file_path)

    assert result == {"imported": 1, "skipped": 0}
    balance = get_order_balances(db_session, "源兴发")[0]
    assert balance["sku"] == "SKU-ZOMBIE-L"


def test_company_export_contains_sku_column(db_session, tmp_path: Path):
    create_order_line(db_session, "源兴发", "小红帽男", "小红帽男", "M", 500)
    upsert_sku_mapping(db_session, "源兴发", "小红帽男", "小红帽男", "M", "SKU-HOOD-M")
    output = tmp_path / "源兴发表.xlsx"

    export_company_workbook(db_session, "源兴发", output)

    ws = load_workbook(output)["订单发货明细"]
    headers = [cell.value for cell in ws[1]]
    assert headers[:12] == ["公司", "产品", "款式", "订单", "下单时间", "尺码", "SKU", "发货明细", "下单数量", "已发数量", "未发数量", "超发数量"]
    assert ws["G2"].value == "SKU-HOOD-M"


def test_create_order_line_auto_resolves_sku_from_mapping(db_session):
    upsert_sku_mapping(db_session, "源兴发", "裁判", "圆领裁判", "S", "SKU-REF-S")

    order = create_order_line(db_session, "源兴发", "裁判", "圆领裁判", "S", 100)

    assert order.sku == "SKU-REF-S"


def test_update_order_line_re_resolves_sku_from_mapping(db_session):
    order = create_order_line(db_session, "源兴发", "裁判", "圆领裁判", "S", 100)
    assert order.sku == ""
    upsert_sku_mapping(db_session, "源兴发", "裁判", "圆领裁判", "M", "SKU-REF-M")

    update_order_line(db_session, order.id, size="M")

    assert order.sku == "SKU-REF-M"

from pathlib import Path

from openpyxl import load_workbook

from app.models import Company, User, WaybillPhoto
from app.services.exports import export_company_workbook
from app.services.orders import create_order_line
from app.services import waybills as waybill_service
from app.services.waybills import build_waybill_file_name, import_waybill_directory, waybill_display_name


PNG_BYTES = (
    b"\x89PNG\r\n\x1a\n\x00\x00\x00\rIHDR\x00\x00\x00\x01\x00\x00\x00\x01"
    b"\x08\x02\x00\x00\x00\x90wS\xde\x00\x00\x00\x0cIDATx\x9cc\xf8\xff\xff?"
    b"\x00\x05\xfe\x02\xfeA\xe2&\xd3\x00\x00\x00\x00IEND\xaeB`\x82"
)


def test_import_waybill_directory_matches_company_folder_and_skips_unknown(db_session, tmp_path: Path, monkeypatch):
    monkeypatch.setattr(waybill_service, "WAYBILL_DIR", tmp_path / "saved_waybills")
    db_session.add_all([
        Company(name="张鹏", is_active=True),
        Company(name="源兴发", is_active=True),
        Company(name="合肥Hoo", is_active=True),
        Company(name="福建", is_active=True),
    ])
    db_session.commit()
    root = tmp_path / "中通快递单"
    (root / "张鹏 蒋先生").mkdir(parents=True)
    (root / "合肥 何玲玲").mkdir(parents=True)
    (root / "杂项").mkdir(parents=True)
    (root / "张鹏 蒋先生" / "7.13.png").write_bytes(PNG_BYTES)
    (root / "合肥 何玲玲" / "7.14.png").write_bytes(PNG_BYTES)
    (root / "杂项" / "未知.png").write_bytes(PNG_BYTES)

    result = import_waybill_directory(db_session, root)

    assert result["imported"] == 2
    assert result["skipped"] == 1
    rows = db_session.query(WaybillPhoto).order_by(WaybillPhoto.company_name).all()
    assert [(row.company_name, row.original_name) for row in rows] == [
        ("合肥Hoo", "2026-07-14_合肥Hoo_面单_01.png"),
        ("张鹏", "2026-07-13_张鹏_面单_01.png"),
    ]


def test_import_waybill_directory_skips_duplicate_file_content(db_session, tmp_path: Path, monkeypatch):
    monkeypatch.setattr(waybill_service, "WAYBILL_DIR", tmp_path / "saved_waybills")
    db_session.add(Company(name="源兴发", is_active=True))
    db_session.commit()
    root = tmp_path / "中通快递单"
    (root / "源兴发 第一次").mkdir(parents=True)
    (root / "源兴发 第二次").mkdir(parents=True)
    (root / "源兴发 第一次" / "7.13.png").write_bytes(PNG_BYTES)
    (root / "源兴发 第二次" / "7.13-copy.png").write_bytes(PNG_BYTES)

    result = import_waybill_directory(db_session, root)

    assert result["imported"] == 1
    assert result["skipped"] == 1
    assert db_session.query(WaybillPhoto).count() == 1


def test_company_export_uses_waybill_number_column_not_photos(db_session, tmp_path: Path):
    user = User(username="admin", display_name="老板", password_hash="x", role="admin", is_active=True)
    db_session.add(user)
    db_session.commit()
    create_order_line(db_session, "源兴发", "小红帽", "小红帽男", "M", 300)
    photo_path = tmp_path / "7.13.png"
    photo_path.write_bytes(PNG_BYTES)
    db_session.add(
        WaybillPhoto(
            company_name="源兴发",
            stored_path=str(photo_path),
            original_name="7.13.png",
            waybill_date="2026-07-13",
            source_path=str(photo_path),
            uploaded_by=user.id,
        )
    )
    db_session.commit()
    output = tmp_path / "源兴发表.xlsx"

    export_company_workbook(db_session, "源兴发", output)

    ws = load_workbook(output)["订单发货明细"]
    headers = [cell.value for cell in ws[1]]
    assert "快递单号" in headers
    assert "快递面单" not in headers
    assert len(ws._images) == 0


def test_waybill_display_name_uses_file_date_before_original_name(db_session, tmp_path: Path):
    photo_path = tmp_path / "7.14.png"
    photo_path.write_bytes(PNG_BYTES)
    photo = WaybillPhoto(company_name="福建", stored_path=str(photo_path), original_name="7.14.png", source_path=str(photo_path))

    assert waybill_display_name(photo) == "2026-07-14 面单"


def test_build_waybill_file_name_uses_date_company_and_index():
    assert build_waybill_file_name("2026-07-13", "源兴发", 2, ".png") == "2026-07-13_源兴发_面单_02.png"

from pathlib import Path

from openpyxl import Workbook
from pypdf import PdfReader
from reportlab.pdfbase import pdfmetrics
from reportlab.pdfbase.cidfonts import UnicodeCIDFont
from reportlab.lib.pagesizes import landscape
from reportlab.pdfgen import canvas

from wr_barcode.core import (
    BarcodeLine,
    generate_barcode_pdf,
    generate_barcode_pdf_for_lines,
    generate_barcode_pdf_from_excel,
    list_contracts_from_excel,
    parse_quantity_table,
    read_contract_quantity_grid,
    read_unshipped_quantity_grid_from_excel,
)


SAMPLE_TABLE = """颜色 尺码 XS S M L XL
黑色2# 25 50 75 55 25
白色39# 60 120 180 125 60
橘色41# 5 10 45 10 5
"""


def test_parse_quantity_table_maps_colors_to_skus_and_adds_extra_quantity():
    lines = parse_quantity_table(SAMPLE_TABLE, extra_each_size=5)

    quantities = {line.sku: line.quantity for line in lines}

    assert quantities["WZY-TS03-BK-XS"] == 30
    assert quantities["WZY-TS03-BK-S"] == 55
    assert quantities["WZY-TS03-WT-M"] == 185
    assert quantities["WZY-TS03-OR-XL"] == 10
    assert len(lines) == 15


def test_parse_quantity_table_accepts_compact_size_quantity_pairs_and_xxl_alias():
    lines = parse_quantity_table("白色S15M30 L25 XL20 XXL10")

    quantities = {(line.color_code, line.size): line.quantity for line in lines}

    assert quantities[("WT", "S")] == 15
    assert quantities[("WT", "M")] == 30
    assert quantities[("WT", "L")] == 25
    assert quantities[("WT", "XL")] == 20
    assert quantities[("WT", "2XL")] == 10
    assert len(lines) == 5


def test_parse_quantity_table_accepts_column_table_with_2xl_header():
    table = """颜色 尺码 XS S M L XL 2XL
黑色 50 150 360 370 300 80
"""

    lines = parse_quantity_table(table)
    quantities = {(line.color_code, line.size): line.quantity for line in lines}

    assert quantities[("BK", "XS")] == 50
    assert quantities[("BK", "XL")] == 300
    assert quantities[("BK", "2XL")] == 80
    assert len(lines) == 6


def test_generate_barcode_pdf_matches_pdf_prefix_and_2xl_size(tmp_path):
    source_pdf = tmp_path / "source_064.pdf"
    output_pdf = tmp_path / "output_064.pdf"
    _write_source_pdf(
        source_pdf,
        [
            "ZCY-DZ-064-WT-S",
            "ZCY-DZ-064-WT-M",
            "ZCY-DZ-064-WT-L",
            "ZCY-DZ-064-WT-XL",
            "ZCY-DZ-064-WT-2XL",
        ],
    )

    result = generate_barcode_pdf(
        source_pdf,
        "白色S1M2 L3 XL4 XXL5",
        output_pdf,
        extra_each_size=1,
    )

    reader = PdfReader(str(output_pdf))
    extracted = [
        (page.extract_text() or "").splitlines()[0].strip()
        for page in reader.pages
    ]
    assert result.total_pages == 20
    assert extracted.count("ZCY-DZ-064-WT-S") == 2
    assert extracted.count("ZCY-DZ-064-WT-2XL") == 6


def test_excel_contract_reader_uses_unshipped_quantity_and_skips_non_positive_rows(tmp_path):
    workbook = tmp_path / "wr.xlsx"
    _write_wr_workbook(workbook)

    contracts = list_contracts_from_excel(workbook)

    assert contracts == ["WR-1 | 豪天 | 064T恤 | 2026-07-07 | 未发 30"]


def test_read_unshipped_quantity_grid_from_total_workbook_groups_selected_contract(tmp_path):
    workbook = tmp_path / "wr.xlsx"
    _write_wr_workbook(workbook)

    grid = read_unshipped_quantity_grid_from_excel(workbook, "WR-1")

    assert grid.sizes == ["S", "M"]
    assert len(grid.rows) == 1
    assert grid.rows[0].color_label == "白色39#"
    assert grid.rows[0].quantities == {"S": 10, "M": 20}
    assert grid.rows[0].total == 30


def test_generate_barcode_pdf_from_excel_matches_chinese_color_text_in_source_pdf(tmp_path):
    workbook = tmp_path / "wr.xlsx"
    source_pdf = tmp_path / "source_064.pdf"
    output_pdf = tmp_path / "output_064.pdf"
    _write_wr_workbook(workbook)
    _write_source_pdf(
        source_pdf,
        [
            "ZCY-DZ-064-WT-S\n白色 S-T 恤 064",
            "ZCY-DZ-064-WT-M\n白色 M-T 恤 064",
            "ZCY-DZ-064-BK-S\n黑色 S-T 恤 064",
        ],
    )

    result = generate_barcode_pdf_from_excel(
        source_pdf=source_pdf,
        excel_path=workbook,
        contract_no="WR-1",
        output_pdf=output_pdf,
        extra_each_size=1,
    )

    reader = PdfReader(str(output_pdf))
    extracted = [
        (page.extract_text() or "").splitlines()[0].strip()
        for page in reader.pages
    ]
    assert result.total_pages == 32
    assert extracted.count("ZCY-DZ-064-WT-S") == 11
    assert extracted.count("ZCY-DZ-064-WT-M") == 21


def test_generate_barcode_pdf_from_excel_uses_color_aliases(tmp_path):
    workbook = tmp_path / "wr_alias.xlsx"
    source_pdf = tmp_path / "source_alias.pdf"
    output_pdf = tmp_path / "output_alias.pdf"
    wb = Workbook()
    ws = wb.active
    ws.title = "订单发货整合"
    ws.append(["合同号", "客户", "款号", "下单日期", "颜色", "尺码", "下单数量", "已发数量", "还差数量"])
    ws.append(["WR-2", "豪天", "064T恤", "2026-07-21", "棕波（94#）", "XL", 100, 30, 70])
    wb.save(workbook)
    _write_source_pdf(source_pdf, ["ZCY-DZ-064-BGBW-XL\n杏底棕波 XL-T 恤 064"])

    result = generate_barcode_pdf_from_excel(source_pdf, workbook, "WR-2", output_pdf)

    reader = PdfReader(str(output_pdf))
    assert result.total_pages == 70
    assert (reader.pages[0].extract_text() or "").splitlines()[0].strip() == "ZCY-DZ-064-BGBW-XL"


def test_read_contract_quantity_grid_finds_raw_contract_color_rows(tmp_path):
    workbook = tmp_path / "raw_contract.xlsx"
    _write_raw_contract_workbook(workbook)

    grid = read_contract_quantity_grid(workbook)

    assert grid.sizes == ["XS", "S", "M", "L", "XL", "2XL"]
    assert grid.rows[0].color_label == "黑色"
    assert grid.rows[0].quantities["XS"] == 50
    assert grid.rows[0].quantities["2XL"] == 80
    assert grid.rows[1].color_label == "灰色"
    assert grid.rows[1].quantities["S"] == 100
    assert grid.rows[1].total == 720


def test_generate_barcode_pdf_for_checked_contract_rows_allows_modified_quantities(tmp_path):
    source_pdf = tmp_path / "source_139.pdf"
    output_pdf = tmp_path / "output_139.pdf"
    _write_source_pdf(
        source_pdf,
        [
            "TM-TS139-GY-S\nTS139 灰色 S",
            "TM-TS139-GY-M\nTS139 灰色 M",
        ],
    )
    lines = [
        BarcodeLine(color_label="灰色", color_code="", size="S", quantity=2),
        BarcodeLine(color_label="灰色", color_code="", size="M", quantity=3),
    ]

    result = generate_barcode_pdf_for_lines(source_pdf, lines, output_pdf)

    reader = PdfReader(str(output_pdf))
    extracted = [(page.extract_text() or "").splitlines()[0].strip() for page in reader.pages]
    assert result.total_pages == 5
    assert extracted.count("TM-TS139-GY-S") == 2
    assert extracted.count("TM-TS139-GY-M") == 3


def test_generate_barcode_pdf_matches_size_first_pdf_labels(tmp_path):
    source_pdf = tmp_path / "source_6128.pdf"
    output_pdf = tmp_path / "output_6128.pdf"
    _write_source_pdf(
        source_pdf,
        [
            "TX6128-BU-S\nS 蓝色女士喇叭下摆 T 恤",
            "TX6128-CM-XS\nXS 咖色女士喇叭下摆 T 恤",
        ],
    )
    lines = [
        BarcodeLine(color_label="蓝色44#", color_code="", size="S", quantity=2),
        BarcodeLine(color_label="咖色58#", color_code="", size="XS", quantity=3),
    ]

    result = generate_barcode_pdf_for_lines(source_pdf, lines, output_pdf)

    reader = PdfReader(str(output_pdf))
    extracted = [(page.extract_text() or "").splitlines()[0].strip() for page in reader.pages]
    assert result.total_pages == 5
    assert extracted.count("TX6128-BU-S") == 2
    assert extracted.count("TX6128-CM-XS") == 3


def test_manual_column_table_allows_colors_matched_by_pdf_text(tmp_path):
    source_pdf = tmp_path / "source_ts01.pdf"
    output_pdf = tmp_path / "output_ts01.pdf"
    _write_source_pdf(
        source_pdf,
        [
            "WZY-TS01-CF-XS\n上⾐ 01 咖⾊ XS",
            "WZY-TS01-CF-S\n上⾐ 01 咖⾊ S",
        ],
    )
    table = """颜色 尺码 XS S M L XL 2XL
咖色17#（定制东盛13129色卡） 10 30 0 0 0 0
"""

    result = generate_barcode_pdf(source_pdf, table, output_pdf)

    reader = PdfReader(str(output_pdf))
    extracted = [(page.extract_text() or "").splitlines()[0].strip() for page in reader.pages]
    assert result.total_pages == 40
    assert extracted.count("WZY-TS01-CF-XS") == 10
    assert extracted.count("WZY-TS01-CF-S") == 30


def test_generate_barcode_pdf_repeats_source_pages_by_requested_quantity(tmp_path):
    source_pdf = tmp_path / "source.pdf"
    output_pdf = tmp_path / "output.pdf"
    _write_source_pdf(
        source_pdf,
        [
            "WZY-TS03-BK-XS",
            "WZY-TS03-BK-S",
            "WZY-TS03-BK-M",
            "WZY-TS03-BK-L",
            "WZY-TS03-BK-XL",
        ],
    )
    table = """颜色 尺码 XS S M L XL
黑色2# 1 2 3 4 5
"""

    result = generate_barcode_pdf(source_pdf, table, output_pdf, extra_each_size=1)

    reader = PdfReader(str(output_pdf))
    extracted = [
        (page.extract_text() or "").splitlines()[0].strip()
        for page in reader.pages
    ]
    assert result.total_pages == 20
    assert result.unique_skus == 5
    assert len(reader.pages) == 20
    assert extracted.count("WZY-TS03-BK-XS") == 2
    assert extracted.count("WZY-TS03-BK-XL") == 6


def _write_source_pdf(path: Path, skus: list[str]) -> None:
    pdfmetrics.registerFont(UnicodeCIDFont("STSong-Light"))
    page_size = landscape((85, 170))
    pdf = canvas.Canvas(str(path), pagesize=page_size)
    for sku in skus:
        for offset, line in enumerate(sku.splitlines()):
            font_name = "STSong-Light" if any(ord(char) > 127 for char in line) else "Helvetica"
            pdf.setFont(font_name, 12)
            pdf.drawString(10, 60 - offset * 16, line)
        pdf.showPage()
    pdf.save()


def _write_wr_workbook(path: Path) -> None:
    wb = Workbook()
    ws = wb.active
    ws.title = "订单发货整合"
    ws.append(["合同号", "客户", "款号", "下单日期", "颜色", "尺码", "下单数量", "已发数量", "还差数量"])
    ws.append(["WR-1", "豪天", "064T恤", "2026-07-07", "白色39#", "S", 20, 10, 10])
    ws.append(["WR-1", "豪天", "064T恤", "2026-07-07", "白色39#", "M", 30, 10, 20])
    ws.append(["WR-1", "豪天", "064T恤", "2026-07-07", "黑色64#", "S", 5, 5, 0])
    ws.append(["WR-1", "豪天", "064T恤", "2026-07-07", "黑色64#", "M", 5, 6, -1])
    wb.save(path)


def _write_raw_contract_workbook(path: Path) -> None:
    wb = Workbook()
    ws = wb.active
    ws.title = "数量"
    ws.append(["产品购销合同"])
    for _ in range(5):
        ws.append([])
    ws.append(["颜色            尺码", "图片", "XS", "S", "M", "L", "XL", "2XL", "总数"])
    ws.append(["黑色", "色卡2#", 50, 150, 360, 370, 300, 80, 1310])
    ws.append(["灰色", "色卡37#", None, 100, 200, 240, 160, 20, 720])
    ws.append(["合计", None, None, None, None, None, None, None, 2030])
    wb.save(path)

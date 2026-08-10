from pathlib import Path

from openpyxl import load_workbook

from app.models import Company, ShipmentLine, ShipmentReport, Spu, User, WaybillRecord
from app.services.exports import export_customer_company_workbook
from app.services.logistics import create_waybill_record, link_reports_to_waybill
from app.services.sales_orders import create_sales_order


def _assert_order_header_once_and_has_order_time(headers):
    assert headers.count("订单") == 1
    order_index = headers.index("订单")
    assert headers[order_index + 1] == "下单时间"


def _setup(db_session):
    admin = User(username="export_admin", display_name="老板", password_hash="x", role="admin", is_active=True)
    company = Company(name="广东茉莉", code="ML", next_order_sequence=1)
    spu = Spu(code="CPYL", product_name="裁判", style_name="圆领裁判")
    db_session.add_all([admin, company, spu])
    db_session.commit()
    formal_order = create_sales_order(
        db_session,
        company.id,
        spu.id,
        "红色",
        "RED",
        "2026-05-22",
        [{"size": "M", "quantity": 100, "customer_sku": "ML-RED-M"}],
        customer_order_no="PO-ML-001",
    )
    order_line = formal_order.lines[0]
    report = ShipmentReport(
        user_id=admin.id,
        order_id=formal_order.id,
        ship_date="2026-07-17",
        company_name="广东茉莉",
        product_name="裁判",
        style_name="圆领裁判",
        status="auto_approved",
    )
    db_session.add(report)
    db_session.flush()
    db_session.add(ShipmentLine(report_id=report.id, order_line_id=order_line.id, size="M", quantity=50))
    db_session.commit()
    waybill = create_waybill_record(db_session, admin.id, "广东茉莉", "2026-07-17", "800209579798", weight_kg=262.3, package_count=9)
    link_reports_to_waybill(db_session, waybill.id, [report.id])
    return formal_order, order_line


def test_balance_projection_exposes_formal_order_number_and_date(db_session):
    from app.services.orders import get_order_balances

    formal_order, _order_line = _setup(db_session)

    row = get_order_balances(db_session, company_name="广东茉莉")[0]

    assert row["system_order_no"] == formal_order.system_order_no
    assert row["formal_order_date"] == "2026-05-22"


def test_customer_export_uses_waybill_number_and_detail_sheet(db_session, tmp_path: Path):
    formal_order, _order_line = _setup(db_session)
    output = tmp_path / "茉莉客户版.xlsx"

    export_customer_company_workbook(db_session, "广东茉莉", output)

    wb = load_workbook(output)
    assert wb.sheetnames == ["客户发货明细", "发货明细"]
    main = wb["客户发货明细"]
    headers = [cell.value for cell in main[1]]
    assert headers[:6] == ["公司", "产品", "款式", "订单", "下单时间", "尺码"]
    _assert_order_header_once_and_has_order_time(headers)
    values = [cell.value for cell in main[2]]
    assert values[3:5] == [formal_order.system_order_no, "2026-05-22"]
    assert values[10] == "800209579798"
    detail = wb["发货明细"]
    detail_headers = [cell.value for cell in detail[1]]
    assert detail_headers[:7] == ["发货日期", "公司", "产品", "款式", "尺码", "数量", "快递单号"]
    assert detail_headers[7:] == ["订单", "下单时间", "客户订单号", "颜色", "SPU", "客户SKU"]
    _assert_order_header_once_and_has_order_time(detail_headers)
    detail_row = [cell.value for cell in detail[2]]
    assert detail_row[0] == "2026-07-17"
    assert detail_row[5] == 50
    assert detail_row[6] == "800209579798"
    assert detail_row[7:] == [formal_order.system_order_no, "2026-05-22", "PO-ML-001", "红色", "CPYL", "ML-RED-M"]


def test_internal_export_shows_waybill_number_not_photos(db_session, tmp_path: Path):
    from app.services.exports import export_company_workbook

    formal_order, _order_line = _setup(db_session)
    output = tmp_path / "茉莉内部版.xlsx"

    export_company_workbook(db_session, "广东茉莉", output)

    wb = load_workbook(output)
    main = wb["订单发货明细"]
    main_headers = [cell.value for cell in main[1]]
    assert main_headers[:6] == ["公司", "产品", "款式", "订单", "下单时间", "尺码"]
    _assert_order_header_once_and_has_order_time(main_headers)
    assert "快递单号" in main_headers
    assert "快递面单" not in main_headers
    main_row = [cell.value for cell in main[2]]
    assert main_row[3:5] == [formal_order.system_order_no, "2026-05-22"]
    waybill_index = main_headers.index("快递单号")
    assert main_row[waybill_index] == "800209579798"
    shipments = wb["发货流水"]
    shipment_headers = [cell.value for cell in shipments[1]]
    _assert_order_header_once_and_has_order_time(shipment_headers)
    order_index = shipment_headers.index("订单")
    shipment_row = [cell.value for cell in shipments[2]]
    assert shipment_row[order_index:order_index + 2] == [formal_order.system_order_no, "2026-05-22"]


def test_total_exports_keep_order_and_add_order_time_on_unshipped_sheets(db_session, tmp_path: Path):
    from app.services.exports import export_customer_total_workbook, export_total_workbook

    formal_order, _order_line = _setup(db_session)
    internal_output = tmp_path / "内部总表.xlsx"
    customer_output = tmp_path / "客户总表.xlsx"

    export_total_workbook(db_session, internal_output)
    export_customer_total_workbook(db_session, customer_output)

    for path in (internal_output, customer_output):
        wb = load_workbook(path)
        ws = wb["未发货明细"]
        headers = [cell.value for cell in ws[1]]
        assert headers[:6] == ["公司", "产品", "款式", "订单", "下单时间", "尺码"]
        _assert_order_header_once_and_has_order_time(headers)
        row = [cell.value for cell in ws[2]]
        assert row[3:5] == [formal_order.system_order_no, "2026-05-22"]


def test_customer_export_leaves_formal_order_fields_blank_when_unbound(db_session, tmp_path: Path):
    from app.services.orders import create_order_line

    create_order_line(db_session, "无绑定公司", "测试产品", "测试款式", "S", 10, order_date="2026-05-01")
    output = tmp_path / "未绑定客户版.xlsx"

    export_customer_company_workbook(db_session, "无绑定公司", output)

    wb = load_workbook(output)
    headers = [cell.value for cell in wb["客户发货明细"][1]]
    _assert_order_header_once_and_has_order_time(headers)
    row = [cell.value for cell in wb["客户发货明细"][2]]
    assert row[3] in (None, "")
    assert row[4] in (None, "")


def test_exports_use_each_historical_line_formal_order_identity(db_session, tmp_path: Path):
    from app.services.exports import export_company_workbook

    admin = User(username="cross_order_export_admin", display_name="老板", password_hash="x", role="admin", is_active=True)
    company = Company(name="跨订单导出客户", code="XDD", next_order_sequence=1)
    spu = Spu(code="KDD", product_name="测试产品", style_name="测试款式")
    db_session.add_all([admin, company, spu])
    db_session.commit()
    first_order = create_sales_order(
        db_session,
        company.id,
        spu.id,
        "红色",
        "RED",
        "2026-01-02",
        [{"size": "S", "quantity": 100, "customer_sku": "SKU-RED-S"}],
        customer_order_no="PO-RED",
    )
    second_order = create_sales_order(
        db_session,
        company.id,
        spu.id,
        "紫色",
        "PURPLE",
        "2026-03-04",
        [{"size": "M", "quantity": 100, "customer_sku": "SKU-PURPLE-M"}],
        customer_order_no="PO-PURPLE",
    )
    historical = ShipmentReport(
        user_id=admin.id,
        order_id=first_order.id,
        ship_date="历史导入",
        company_name=company.name,
        product_name=spu.product_name,
        style_name=spu.style_name,
        status="auto_approved",
    )
    db_session.add(historical)
    db_session.flush()
    db_session.add_all(
        [
            ShipmentLine(report_id=historical.id, order_line_id=first_order.lines[0].id, size="S", quantity=5),
            ShipmentLine(report_id=historical.id, order_line_id=second_order.lines[0].id, size="M", quantity=6),
        ]
    )
    db_session.commit()

    customer_output = tmp_path / "跨订单客户版.xlsx"
    internal_output = tmp_path / "跨订单内部版.xlsx"
    export_customer_company_workbook(db_session, company.name, customer_output)
    export_company_workbook(db_session, company.name, internal_output)

    for ws in (load_workbook(customer_output)["发货明细"], load_workbook(internal_output)["发货流水"]):
        headers = [cell.value for cell in ws[1]]
        _assert_order_header_once_and_has_order_time(headers)
        rows_by_size = {
            row[headers.index("尺码")]: row
            for row in ws.iter_rows(min_row=2, values_only=True)
        }
        order_index = headers.index("订单")
        assert rows_by_size["S"][order_index:order_index + 2] == (first_order.system_order_no, "2026-01-02")
        assert rows_by_size["M"][order_index:order_index + 2] == (second_order.system_order_no, "2026-03-04")

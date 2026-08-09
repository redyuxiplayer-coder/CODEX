from pathlib import Path

from openpyxl import load_workbook

from app.api_v1 import _report_dict
from app.models import Company, ShipmentLine, ShipmentReport, Spu, User
from app.services.exports import export_daily_shipments_workbook
from app.services.logistics import list_unlinked_reports
from app.services.sales_orders import create_sales_order
from app.services.shipments import submit_shipment_report


def _formal_shipment(db_session):
    company = Company(name="源兴发", code="YXF", next_order_sequence=1)
    spu = Spu(code="JS", product_name="啦啦队", style_name="僵尸啦啦队")
    admin = User(username="display_admin", display_name="老板", password_hash="x", role="admin", is_active=True)
    db_session.add_all([company, spu, admin])
    db_session.commit()
    order = create_sales_order(
        db_session,
        company.id,
        spu.id,
        "红色",
        "RED",
        "2026-08-09",
        [{"size": "S", "quantity": 100, "customer_sku": "FZB1209001-01-red-S"}],
        customer_order_no="PO-7788",
    )
    report = submit_shipment_report(
        db_session,
        admin.id,
        "2026-08-10",
        "",
        "",
        "",
        [{"size": "S", "quantity": 20, "order_line_id": order.lines[0].id}],
        order_id=order.id,
    )
    return order, report


def test_shipment_api_payload_includes_formal_order_identity(db_session):
    order, report = _formal_shipment(db_session)

    payload = _report_dict(report)

    assert payload["order_id"] == order.id
    assert payload["system_order_no"] == "YXF-00001-JS-RED"
    assert payload["customer_order_no"] == "PO-7788"
    assert payload["order_date"] == "2026-08-09"
    assert payload["color_name"] == "红色"
    assert payload["spu_code"] == "JS"
    assert payload["lines"][0]["customer_sku"] == "FZB1209001-01-red-S"


def test_historical_cross_order_report_lists_each_line_order_identity(db_session):
    first_order, _report = _formal_shipment(db_session)
    second_order = create_sales_order(
        db_session,
        first_order.company_id,
        first_order.spu_id,
        "紫色",
        "PURPLE",
        "2026-08-11",
        [{"size": "M", "quantity": 100, "customer_sku": ""}],
    )
    historical = ShipmentReport(
        user_id=_report.user_id,
        ship_date="2026-08-12",
        company_name="源兴发",
        product_name="啦啦队",
        style_name="僵尸啦啦队",
        status="auto_approved",
    )
    db_session.add(historical)
    db_session.flush()
    db_session.add_all(
        [
            ShipmentLine(report_id=historical.id, order_line_id=first_order.lines[0].id, size="S", quantity=5),
            ShipmentLine(report_id=historical.id, order_line_id=second_order.lines[0].id, size="M", quantity=6),
        ]
    )
    db_session.commit()

    payload = _report_dict(historical)

    assert payload["order_id"] is None
    assert payload["system_order_no"] == "YXF-00001-JS-RED / YXF-00002-JS-PURPLE"
    assert [line["system_order_no"] for line in payload["lines"]] == [
        "YXF-00001-JS-RED",
        "YXF-00002-JS-PURPLE",
    ]
    assert [line["order_date"] for line in payload["lines"]] == ["2026-08-09", "2026-08-11"]


def test_unlinked_logistics_row_includes_system_order_number(db_session):
    _order, report = _formal_shipment(db_session)

    rows = list_unlinked_reports(db_session)

    assert rows[0]["id"] == report.id
    assert rows[0]["system_order_no"] == "YXF-00001-JS-RED"
    assert rows[0]["order_date"] == "2026-08-09"


def test_daily_shipment_export_identifies_order_and_customer_sku(db_session, tmp_path: Path):
    _formal_shipment(db_session)
    output = tmp_path / "daily.xlsx"

    export_daily_shipments_workbook(db_session, "2026-08-10", output)

    ws = load_workbook(output)["发货流水"]
    headers = [cell.value for cell in ws[1]]
    row = {headers[index]: cell.value for index, cell in enumerate(ws[2])}
    assert row["系统订单号"] == "YXF-00001-JS-RED"
    assert row["客户订单号"] == "PO-7788"
    assert row["下单日期"] == "2026-08-09"
    assert row["颜色"] == "红色"
    assert row["SPU"] == "JS"
    assert row["客户SKU"] == "FZB1209001-01-red-S"

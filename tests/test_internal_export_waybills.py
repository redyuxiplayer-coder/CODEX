from pathlib import Path

from openpyxl import load_workbook

from app.models import ShipmentReport, User
from app.services.exports import (
    export_company_workbook,
    export_customer_company_workbook,
    export_customer_total_workbook,
    export_total_workbook,
)
from app.services.logistics import create_waybill_record, link_reports_to_waybill


def _create_admin(db_session) -> User:
    admin = User(
        username="waybill_export_admin",
        display_name="老板",
        password_hash="x",
        role="admin",
        is_active=True,
    )
    db_session.add(admin)
    db_session.commit()
    return admin


def _add_report(db_session, admin: User, company_name: str, ship_date: str) -> ShipmentReport:
    report = ShipmentReport(
        user_id=admin.id,
        ship_date=ship_date,
        company_name=company_name,
        product_name="测试产品",
        style_name="测试款式",
        status="auto_approved",
    )
    db_session.add(report)
    db_session.commit()
    return report


def test_internal_total_export_adds_sorted_unique_waybill_sheet(db_session, tmp_path: Path):
    admin = _create_admin(db_session)
    zero_weight = create_waybill_record(
        db_session, admin.id, "甲公司", "2026-07-02", "WB-200", weight_kg=0
    )
    create_waybill_record(
        db_session, admin.id, "乙公司", "2026-07-01", "WB-300", weight_kg=12.5
    )
    linked = create_waybill_record(
        db_session, admin.id, "甲公司", "2026-07-01", "WB-100", weight_kg=8
    )
    first_report = _add_report(db_session, admin, "甲公司", "2026-07-01")
    second_report = _add_report(db_session, admin, "甲公司", "2026-07-01")
    link_reports_to_waybill(db_session, linked.id, [first_report.id, second_report.id])

    output = export_total_workbook(db_session, tmp_path / "内部总表.xlsx")

    workbook = load_workbook(output)
    assert workbook.sheetnames == ["订单发货明细", "发货流水", "未发货明细", "快递记录"]
    sheet = workbook["快递记录"]
    assert list(sheet.values) == [
        ("快递单号", "重量(kg)", "发货日期"),
        ("WB-100", 8, "2026-07-01"),
        ("WB-300", 12.5, "2026-07-01"),
        ("WB-200", None, "2026-07-02"),
    ]
    assert sheet.freeze_panes == "A2"
    assert sheet["A1"].font.bold is True
    assert zero_weight.weight_kg == 0


def test_internal_company_export_filters_waybills_and_keeps_empty_sheet(db_session, tmp_path: Path):
    admin = _create_admin(db_session)
    create_waybill_record(
        db_session, admin.id, "甲公司", "2026-07-02", "A-002", weight_kg=6.5
    )
    create_waybill_record(
        db_session, admin.id, "甲公司", "2026-07-01", "A-001", weight_kg=5
    )
    create_waybill_record(
        db_session, admin.id, "乙公司", "2026-07-01", "B-001", weight_kg=9
    )

    company_output = export_company_workbook(db_session, "甲公司", tmp_path / "甲公司内部版.xlsx")
    empty_output = export_company_workbook(db_session, "无快递公司", tmp_path / "空内部版.xlsx")

    company_workbook = load_workbook(company_output)
    assert company_workbook.sheetnames == ["订单发货明细", "发货流水", "快递记录"]
    assert list(company_workbook["快递记录"].values) == [
        ("快递单号", "重量(kg)", "发货日期"),
        ("A-001", 5, "2026-07-01"),
        ("A-002", 6.5, "2026-07-02"),
    ]
    assert list(load_workbook(empty_output)["快递记录"].values) == [
        ("快递单号", "重量(kg)", "发货日期"),
    ]


def test_customer_exports_do_not_add_waybill_sheet(db_session, tmp_path: Path):
    admin = _create_admin(db_session)
    create_waybill_record(
        db_session, admin.id, "甲公司", "2026-07-01", "A-001", weight_kg=5
    )

    total_output = export_customer_total_workbook(db_session, tmp_path / "客户总表.xlsx")
    company_output = export_customer_company_workbook(
        db_session, "甲公司", tmp_path / "甲公司客户版.xlsx"
    )

    assert load_workbook(total_output).sheetnames == ["客户发货明细", "未发货明细", "发货明细"]
    assert load_workbook(company_output).sheetnames == ["客户发货明细", "发货明细"]

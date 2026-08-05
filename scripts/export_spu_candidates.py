"""导出 SPU 分类候选表：列出所有「公司+产品+款式」组合，供人工归类。

用法（本地默认连 SQLite）：
    python scripts/export_spu_candidates.py [输出路径]

服务器用法（连生产库）：
    先确认 app 启动时的 SUPABASE_DATABASE_URL 环境变量，然后：
    SUPABASE_DATABASE_URL=... python scripts/export_spu_candidates.py /home/ubuntu/zy-shipping/data/exports/SPU分类表.xlsx
"""

import os
import sys
from collections import defaultdict
from pathlib import Path

REPO_ROOT = Path(__file__).resolve().parent.parent
sys.path.insert(0, str(REPO_ROOT))

from openpyxl import Workbook
from openpyxl.styles import Alignment, Font, PatternFill
from openpyxl.utils import get_column_letter

from app.config import EXPORT_DIR
from app.db import SessionLocal
from app.models import Company, OrderLine, SkuMapping


def main() -> int:
    output = Path(sys.argv[1]) if len(sys.argv) > 1 else EXPORT_DIR / "SPU分类表.xlsx"
    output.parent.mkdir(parents=True, exist_ok=True)

    session = SessionLocal()
    try:
        rows = (
            session.query(
                Company.name.label("company"),
                OrderLine.product_name.label("product"),
                OrderLine.style_name.label("style"),
            )
            .join(Company, Company.id == OrderLine.company_id)
            .filter(
                OrderLine.is_active.is_(True),
                Company.is_active.is_(True),
                Company.name != "演示",
            )
            .distinct()
            .order_by(Company.name, OrderLine.product_name, OrderLine.style_name)
            .all()
        )
        combos = [(row.company, row.product, row.style) for row in rows]

        sizes_by_combo: dict[tuple[str, str, str], set[str]] = defaultdict(set)
        qty_by_combo: dict[tuple[str, str, str], int] = defaultdict(int)
        for company, product, style, size, quantity in session.query(
            Company.name,
            OrderLine.product_name,
            OrderLine.style_name,
            OrderLine.size,
            OrderLine.quantity,
        ).join(Company, Company.id == OrderLine.company_id).filter(
            OrderLine.is_active.is_(True),
            Company.is_active.is_(True),
            Company.name != "演示",
        ).all():
            key = (company, product, style)
            sizes_by_combo[key].add(size or "")
            qty_by_combo[key] += int(quantity or 0)

        sku_by_combo: dict[tuple[str, str, str], set[str]] = defaultdict(set)
        for company, product, style, sku in session.query(
            SkuMapping.company_name,
            SkuMapping.product_name,
            SkuMapping.style_name,
            SkuMapping.sku,
        ).all():
            if sku:
                sku_by_combo[(company, product, style)].add(sku)
    finally:
        session.close()

    wb = Workbook()
    ws = wb.active
    ws.title = "SPU分类表"
    headers = ["公司", "产品", "款式", "尺码", "下单合计", "客户SKU", "SPU码(待填)", "SPU名称(待填)", "备注"]
    ws.append(headers)
    header_fill = PatternFill("solid", fgColor="E8EEF7")
    for cell in ws[1]:
        cell.fill = header_fill
        cell.font = Font(bold=True)
        cell.alignment = Alignment(horizontal="center", vertical="center")

    def size_sort(value: str) -> tuple[int, str]:
        order = {"S": 0, "M": 1, "L": 2, "XL": 3, "XXL": 4}
        return (order.get(value.upper(), 99), value)

    for combo in combos:
        company, product, style = combo
        sizes = ", ".join(sorted(sizes_by_combo.get(combo, set()), key=size_sort))
        skus = ", ".join(sorted(sku_by_combo.get(combo, set())))
        ws.append([company, product, style, sizes, qty_by_combo.get(combo, 0), skus, "", "", ""])

    widths = [12, 18, 26, 20, 10, 30, 16, 20, 24]
    for index, width in enumerate(widths, start=1):
        ws.column_dimensions[get_column_letter(index)].width = width
    ws.freeze_panes = "A2"
    wb.save(output)
    print(f"已生成 {output}（{len(combos)} 个组合）")
    return 0


if __name__ == "__main__":
    sys.exit(main())

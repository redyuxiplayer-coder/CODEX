from __future__ import annotations

from dataclasses import dataclass
from pathlib import Path
import re

from openpyxl import load_workbook
from pypdf import PdfReader, PdfWriter


SIZE_ORDER = ("XS", "S", "M", "L", "XL")
SIZE_ORDER_WITH_2XL = ("XS", "S", "M", "L", "XL", "2XL")
SIZE_ALIASES = {
    "XS": "XS",
    "S": "S",
    "M": "M",
    "L": "L",
    "XL": "XL",
    "XXL": "2XL",
    "2XL": "2XL",
}
SIZE_PATTERN = r"(?:XXL|2XL|XL|XS|S|M|L)"

COLOR_SKU_CODES = {
    "黑色": "BK",
    "黑": "BK",
    "白色": "WT",
    "白": "WT",
    "深蓝": "DBL",
    "粉色": "PK",
    "粉": "PK",
    "孔雀绿": "PCG",
    "蓝色": "BL",
    "蓝": "BL",
    "黄色": "YE",
    "黄": "YE",
    "灰色": "GY",
    "灰": "GY",
    "橘色": "OR",
    "橘": "OR",
}

COLOR_LABEL_ALIASES = {
    "棕波": "杏底棕波",
}


@dataclass(frozen=True)
class BarcodeLine:
    color_label: str
    color_code: str
    size: str
    quantity: int

    @property
    def sku(self) -> str:
        return f"WZY-TS03-{self.color_code}-{self.size}"


@dataclass(frozen=True)
class GenerationResult:
    output_pdf: Path
    manifest_path: Path
    total_pages: int
    unique_skus: int


@dataclass(frozen=True)
class ContractSummary:
    contract_no: str
    customer: str
    style_no: str
    order_date: str
    unshipped_quantity: int

    @property
    def display(self) -> str:
        return (
            f"{self.contract_no} | {self.customer} | {self.style_no} | "
            f"{self.order_date} | 未发 {self.unshipped_quantity}"
        )


@dataclass(frozen=True)
class QuantityGridRow:
    color_label: str
    quantities: dict[str, int]

    @property
    def total(self) -> int:
        return sum(self.quantities.values())


@dataclass(frozen=True)
class QuantityGrid:
    sizes: list[str]
    rows: list[QuantityGridRow]


def parse_quantity_table(text: str, extra_each_size: int = 0) -> list[BarcodeLine]:
    if extra_each_size < 0:
        raise ValueError("每个尺码额外数量不能小于 0")

    lines: list[BarcodeLine] = []
    for raw_line in text.splitlines():
        row = raw_line.strip()
        if not row:
            continue
        parts = re.split(r"\s+", row)
        if _is_header(parts):
            continue
        if len(parts) in (6, 7) and all(part.isdigit() for part in parts[1:]):
            lines.extend(_parse_column_row(parts[0], parts[1:], row, extra_each_size))
            continue

        compact_lines = _parse_compact_row(row, extra_each_size)
        if compact_lines:
            lines.extend(compact_lines)
            continue

        raise ValueError(f"这一行格式不对：{row}")

    if not lines:
        raise ValueError("没有识别到颜色尺码数量")
    return lines


def generate_barcode_pdf(
    source_pdf: str | Path,
    table_text: str,
    output_pdf: str | Path,
    extra_each_size: int = 0,
) -> GenerationResult:
    source = Path(source_pdf)
    output = Path(output_pdf)
    if not source.exists():
        raise FileNotFoundError(f"找不到源条码 PDF：{source}")

    requests = parse_quantity_table(table_text, extra_each_size)
    return generate_barcode_pdf_for_lines(source, requests, output)


def generate_barcode_pdf_for_lines(
    source_pdf: str | Path,
    requests: list[BarcodeLine],
    output_pdf: str | Path,
) -> GenerationResult:
    source = Path(source_pdf)
    output = Path(output_pdf)
    if not source.exists():
        raise FileNotFoundError(f"找不到源条码 PDF：{source}")

    reader = PdfReader(str(source))
    page_by_sku, page_by_color_size = _build_page_indexes(reader)
    missing = [
        _line_display(line)
        for line in requests
        if line.quantity > 0 and _sku_for_line(line, page_by_sku, page_by_color_size) is None
    ]
    if missing:
        raise ValueError("源 PDF 里缺少这些条码：" + "，".join(missing))

    output.parent.mkdir(parents=True, exist_ok=True)
    writer = PdfWriter()
    manifest_lines: list[str] = []
    actual_skus: set[str] = set()
    for line in requests:
        if line.quantity <= 0:
            continue
        sku = _sku_for_line(line, page_by_sku, page_by_color_size)
        if sku is None:
            raise ValueError(f"源 PDF 里缺少这个条码：{line.sku}")
        actual_skus.add(sku)
        page_index = page_by_sku[sku]
        for _ in range(line.quantity):
            writer.add_page(reader.pages[page_index])
        manifest_lines.append(f"{line.color_label} {line.size}: {line.quantity} 张 ({sku})")

    with output.open("wb") as stream:
        writer.write(stream)

    manifest_path = output.with_suffix(".txt")
    manifest_path.write_text(
        "\n".join(manifest_lines) + f"\n合计: {len(writer.pages)} 张\n",
        encoding="utf-8",
    )
    return GenerationResult(
        output_pdf=output,
        manifest_path=manifest_path,
        total_pages=len(writer.pages),
        unique_skus=len(actual_skus),
    )


def list_contracts_from_excel(excel_path: str | Path) -> list[str]:
    summaries = _read_contract_summaries(excel_path)
    return [summary.display for summary in summaries if summary.unshipped_quantity > 0]


def generate_barcode_pdf_from_excel(
    source_pdf: str | Path,
    excel_path: str | Path,
    contract_no: str,
    output_pdf: str | Path,
    extra_each_size: int = 0,
) -> GenerationResult:
    lines = _read_unshipped_lines_from_excel(excel_path, contract_no, extra_each_size)
    if not lines:
        raise ValueError(f"合同 {contract_no} 没有未发货数量")
    return generate_barcode_pdf_for_lines(source_pdf, lines, output_pdf)


def read_unshipped_quantity_grid_from_excel(excel_path: str | Path, contract_no: str) -> QuantityGrid:
    lines = _read_unshipped_lines_from_excel(excel_path, contract_no, extra_each_size=0)
    if not lines:
        raise ValueError(f"合同 {contract_no} 没有未发货数量")

    sizes = [size for size in SIZE_ORDER_WITH_2XL if any(line.size == size for line in lines)]
    grouped: dict[str, dict[str, int]] = {}
    for line in lines:
        grouped.setdefault(line.color_label, {})[line.size] = (
            grouped.setdefault(line.color_label, {}).get(line.size, 0) + line.quantity
        )

    rows = [
        QuantityGridRow(
            color_label=color_label,
            quantities={size: quantities.get(size, 0) for size in sizes},
        )
        for color_label, quantities in grouped.items()
    ]
    return QuantityGrid(sizes=sizes, rows=rows)


def read_contract_quantity_grid(excel_path: str | Path) -> QuantityGrid:
    path = Path(excel_path)
    if not path.exists():
        raise FileNotFoundError(f"找不到合同文件：{path}")

    workbook = load_workbook(path, read_only=True, data_only=True)
    sheet_name = "数量" if "数量" in workbook.sheetnames else workbook.sheetnames[0]
    ws = workbook[sheet_name]
    header_row, size_columns = _find_quantity_header(ws)
    rows: list[QuantityGridRow] = []

    for row_index in range(header_row + 1, ws.max_row + 1):
        color_value = ws.cell(row_index, 1).value
        color_label = str(color_value).strip() if color_value is not None else ""
        if not color_label or color_label in {"合计", "总计"}:
            if rows:
                break
            continue
        quantities: dict[str, int] = {}
        for size, column_index in size_columns:
            quantities[size] = _as_int(ws.cell(row_index, column_index).value)
        if any(quantity > 0 for quantity in quantities.values()):
            rows.append(QuantityGridRow(color_label=color_label, quantities=quantities))

    if not rows:
        raise ValueError("没有在合同数量页识别到颜色尺码数量")

    return QuantityGrid(sizes=[size for size, _ in size_columns], rows=rows)


def _is_header(parts: list[str]) -> bool:
    joined = " ".join(parts).upper()
    return "颜色" in joined and "XS" in joined and "XL" in joined


def _parse_quantity(value: str, row: str) -> int:
    try:
        quantity = int(value)
    except ValueError as exc:
        raise ValueError(f"这一行有不是整数的数量：{row}") from exc
    if quantity < 0:
        raise ValueError(f"数量不能小于 0：{row}")
    return quantity


def _parse_column_row(
    color_label: str,
    quantity_values: list[str],
    row: str,
    extra_each_size: int,
) -> list[BarcodeLine]:
    color_code = _try_color_code_for(color_label)
    sizes = SIZE_ORDER_WITH_2XL if len(quantity_values) == 6 else SIZE_ORDER
    lines: list[BarcodeLine] = []
    for size, value in zip(sizes, quantity_values):
        lines.append(
            BarcodeLine(
                color_label=color_label,
                color_code=color_code,
                size=size,
                quantity=_parse_quantity(value, row) + extra_each_size,
            )
        )
    return lines


def _parse_compact_row(row: str, extra_each_size: int) -> list[BarcodeLine]:
    compact = re.sub(r"\s+", "", row).upper()
    first_size = re.search(SIZE_PATTERN + r"\d+", compact)
    if not first_size:
        return []

    color_label = row[: first_size.start()].strip()
    if not color_label:
        return []

    color_code = _try_color_code_for(color_label)
    tail = compact[first_size.start() :]
    pairs = list(re.finditer(rf"({SIZE_PATTERN})(\d+)", tail))
    if not pairs or "".join(pair.group(0) for pair in pairs) != tail:
        return []

    return [
        BarcodeLine(
            color_label=color_label,
            color_code=color_code,
            size=SIZE_ALIASES[pair.group(1)],
            quantity=_parse_quantity(pair.group(2), row) + extra_each_size,
        )
        for pair in pairs
    ]


def _color_code_for(color_label: str) -> str:
    color_name = re.sub(r"\d+#?$", "", color_label).strip()
    for name, code in sorted(COLOR_SKU_CODES.items(), key=lambda item: len(item[0]), reverse=True):
        if name in color_name:
            return code
    raise ValueError(f"不认识这个颜色：{color_label}")


def _try_color_code_for(color_label: str) -> str:
    try:
        return _color_code_for(color_label)
    except ValueError:
        return ""


def _build_page_indexes(reader: PdfReader) -> tuple[dict[str, int], dict[tuple[str, str], str]]:
    page_by_sku: dict[str, int] = {}
    page_by_color_size: dict[tuple[str, str], str] = {}
    for index, page in enumerate(reader.pages):
        text = page.extract_text() or ""
        match = re.search(rf"[A-Z0-9]+(?:-[A-Z0-9]+)*-[A-Z]+-{SIZE_PATTERN}", text)
        if match:
            sku = match.group(0)
            page_by_sku[sku] = index
            color_size = _color_size_from_page_text(text)
            if color_size:
                page_by_color_size[color_size] = sku
    return page_by_sku, page_by_color_size


def _build_page_index(reader: PdfReader) -> dict[str, int]:
    page_by_sku, _ = _build_page_indexes(reader)
    return page_by_sku


def _color_size_from_page_text(text: str) -> tuple[str, str] | None:
    lines = [line.strip() for line in text.splitlines() if line.strip()]
    for line in lines[1:] + lines[:1]:
        normalized_line = _normalize_pdf_text(line)
        size_first = re.search(rf"^({SIZE_PATTERN})\s+(.+)$", normalized_line.upper())
        if size_first:
            size = SIZE_ALIASES[size_first.group(1)]
            color = _normalize_color_label(_strip_product_description(size_first.group(2)))
            if color:
                return color, size

        color_first = re.search(rf"(?:[A-Z0-9-]+\s+)?(.+?)\s+({SIZE_PATTERN})(?:-|$|\s)", normalized_line.upper())
        if color_first:
            color = _normalize_color_label(_strip_product_description(color_first.group(1)))
            size = SIZE_ALIASES[color_first.group(2)]
            if color:
                return color, size
    return None


def _sku_for_line(
    line: BarcodeLine,
    page_by_sku: dict[str, int],
    page_by_color_size: dict[tuple[str, str], str] | None = None,
) -> str | None:
    if line.sku in page_by_sku:
        return line.sku

    if line.color_code:
        suffix = f"-{line.color_code}-{line.size}"
        matches = [sku for sku in page_by_sku if sku.endswith(suffix)]
        if len(matches) == 1:
            return matches[0]

    if page_by_color_size:
        color_key = _normalize_color_label(line.color_label)
        sku = page_by_color_size.get((color_key, line.size))
        if sku:
            return sku
    return None


def _read_contract_summaries(excel_path: str | Path) -> list[ContractSummary]:
    path = Path(excel_path)
    if not path.exists():
        raise FileNotFoundError(f"找不到 WR 总表：{path}")
    ws = load_workbook(path, read_only=True, data_only=True)["订单发货整合"]
    headers = _headers(ws)
    grouped: dict[str, ContractSummary] = {}
    totals: dict[str, int] = {}
    for row in ws.iter_rows(min_row=2, values_only=True):
        record = _row_record(headers, row)
        contract_no = str(record["合同号"]).strip()
        unshipped = _as_int(record["还差数量"])
        totals[contract_no] = totals.get(contract_no, 0) + max(unshipped, 0)
        if contract_no not in grouped:
            grouped[contract_no] = ContractSummary(
                contract_no=contract_no,
                customer=str(record["客户"]).strip(),
                style_no=str(record["款号"]).strip(),
                order_date=str(record["下单日期"]).strip(),
                unshipped_quantity=0,
            )

    return [
        ContractSummary(
            contract_no=item.contract_no,
            customer=item.customer,
            style_no=item.style_no,
            order_date=item.order_date,
            unshipped_quantity=totals.get(item.contract_no, 0),
        )
        for item in grouped.values()
    ]


def _read_unshipped_lines_from_excel(
    excel_path: str | Path,
    contract_no: str,
    extra_each_size: int,
) -> list[BarcodeLine]:
    if extra_each_size < 0:
        raise ValueError("每个尺码额外数量不能小于 0")
    path = Path(excel_path)
    if not path.exists():
        raise FileNotFoundError(f"找不到 WR 总表：{path}")
    ws = load_workbook(path, read_only=True, data_only=True)["订单发货整合"]
    headers = _headers(ws)
    totals: dict[tuple[str, str], int] = {}
    labels: dict[tuple[str, str], str] = {}
    for row in ws.iter_rows(min_row=2, values_only=True):
        record = _row_record(headers, row)
        if str(record["合同号"]).strip() != contract_no:
            continue
        unshipped = _as_int(record["还差数量"])
        if unshipped <= 0:
            continue
        color_label = str(record["颜色"]).strip()
        size = SIZE_ALIASES.get(str(record["尺码"]).strip().upper())
        if not size:
            raise ValueError(f"不认识这个尺码：{record['尺码']}")
        key = (color_label, size)
        totals[key] = totals.get(key, 0) + unshipped
        labels[key] = color_label

    return [
        BarcodeLine(
            color_label=color_label,
            color_code=_try_color_code_for(color_label),
            size=size,
            quantity=quantity + extra_each_size,
        )
        for (color_label, size), quantity in totals.items()
    ]


def _headers(ws) -> dict[str, int]:
    row = next(ws.iter_rows(min_row=1, max_row=1, values_only=True))
    headers = {str(value).strip(): index for index, value in enumerate(row)}
    required = {"合同号", "客户", "款号", "下单日期", "颜色", "尺码", "还差数量"}
    missing = required - set(headers)
    if missing:
        raise ValueError("WR 总表缺少这些列：" + "，".join(sorted(missing)))
    return headers


def _find_quantity_header(ws) -> tuple[int, list[tuple[str, int]]]:
    for row in ws.iter_rows(values_only=False):
        size_columns: list[tuple[str, int]] = []
        has_color_header = False
        for cell in row:
            value = str(cell.value).strip().upper() if cell.value is not None else ""
            if "颜色" in value and "尺码" in value:
                has_color_header = True
            if value in SIZE_ALIASES:
                size_columns.append((SIZE_ALIASES[value], cell.column))
        if has_color_header and size_columns:
            return row[0].row, size_columns
    raise ValueError("没有找到合同里的颜色尺码表头")


def _row_record(headers: dict[str, int], row: tuple) -> dict[str, object]:
    return {name: row[index] for name, index in headers.items()}


def _as_int(value: object) -> int:
    if value is None:
        return 0
    return int(value)


def _normalize_color_label(value: str) -> str:
    normalized = str(value).strip()
    normalized = _normalize_pdf_text(normalized)
    normalized = re.sub(r"[（(].*?[）)]", "", normalized)
    normalized = normalized.replace("波点", "波")
    normalized = normalized.replace("色底", "底")
    normalized = normalized.replace("色", "")
    normalized = re.sub(r"\d+#?$", "", normalized)
    normalized = normalized.strip()
    return COLOR_LABEL_ALIASES.get(normalized, normalized)


def _normalize_pdf_text(value: str) -> str:
    return (
        str(value)
        .replace("⽩", "白")
        .replace("⿊", "黑")
        .replace("⻩", "黄")
        .replace("⾊", "色")
        .replace("⾐", "衣")
        .replace("⼥", "女")
        .replace("⼠", "士")
    )


def _strip_product_description(value: str) -> str:
    cleaned = str(value).strip()
    cleaned = re.sub(r"^上衣\s*\d+\s*", "", cleaned, flags=re.IGNORECASE)
    for marker in ("女士", "男士", "喇叭", "下摆", "T 恤", "T恤", "短袖", "上衣", "外套"):
        index = cleaned.find(marker)
        if index > 0:
            cleaned = cleaned[:index]
            break
    return cleaned.strip()


def _line_display(line: BarcodeLine) -> str:
    if line.color_code:
        return line.sku
    return f"{line.color_label} {line.size}"

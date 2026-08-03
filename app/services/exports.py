from pathlib import Path

from openpyxl import Workbook
from openpyxl.drawing.image import Image as ExcelImage
from openpyxl.styles import Alignment, Font, PatternFill
from openpyxl.utils import get_column_letter
from sqlalchemy.orm import Session

from app.models import ShipmentReport
from app.services.aliases import canonical_item
from app.services.orders import get_order_balances
from app.services.waybills import get_waybill_photos, waybill_display_name


def style_sheet(ws) -> None:
    header_fill = PatternFill("solid", fgColor="D9EAF7")
    for cell in ws[1]:
        cell.font = Font(bold=True)
        cell.fill = header_fill
        cell.alignment = Alignment(horizontal="center", vertical="center")
    for row in ws.iter_rows():
        for cell in row:
            cell.alignment = Alignment(horizontal="center", vertical="center", wrap_text=True)
    ws.freeze_panes = "A2"
    for col in range(1, ws.max_column + 1):
        max_len = 10
        for cell in ws[get_column_letter(col)]:
            if cell.value is not None:
                max_len = max(max_len, min(30, len(str(cell.value)) + 2))
        ws.column_dimensions[get_column_letter(col)].width = max_len


def _shipment_detail_text(report: ShipmentReport, size: str, quantity: int) -> str:
    if report.ship_date == "历史导入" and report.note:
        return report.note
    return f"{report.ship_date} 发 {quantity}件"


def _order_label(row: dict) -> str:
    label = str(row.get("order_ref") or row.get("order_date") or "本单").strip()
    if label.endswith("单") or label == "本单":
        return label
    return f"{label}订单"


def _split_shipment_detail_text(report: ShipmentReport, total_quantity: int, allocations: list[tuple[dict, int]]) -> str:
    if report.ship_date == "历史导入" and report.note:
        return report.note
    if len(allocations) <= 1:
        return _shipment_detail_text(report, "", allocations[0][1])
    parts = "，".join(f"{quantity}件发{_order_label(row)}" for row, quantity in allocations)
    return f"{report.ship_date} 发 {total_quantity}件，其中{parts}"


def _balance_detail_key(row: dict):
    return row.get("order_id") or (row["company"], row["product"], row["style"], row["size"], row.get("order_ref", ""))


def shipment_details_by_balance_row(session: Session, balances: list[dict]) -> dict[object, str]:
    details: dict[object, list[str]] = {}
    remaining_capacity: dict[object, int] = {}
    rows_by_base_key: dict[tuple[str, str, str, str], list[dict]] = {}
    rows_by_order_id: dict[int, dict] = {}
    for row in balances:
        key = _balance_detail_key(row)
        remaining_capacity[key] = int(row.get("shipped") or 0)
        base_key = (row["company"], row["product"], row["style"], row["size"])
        rows_by_base_key.setdefault(base_key, []).append(row)
        for order_id in row.get("order_ids", [row.get("order_id")]):
            if order_id:
                rows_by_order_id[int(order_id)] = row

    reports = (
        session.query(ShipmentReport)
        .filter(ShipmentReport.status.in_(("auto_approved", "approved_after_edit")))
        .order_by(ShipmentReport.ship_date, ShipmentReport.created_at)
        .all()
    )
    for report in reports:
        for line in report.lines:
            if line.order_line_id:
                row = rows_by_order_id.get(int(line.order_line_id))
                if row:
                    row_key = _balance_detail_key(row)
                    details.setdefault(row_key, []).append(_shipment_detail_text(report, line.size, int(line.quantity or 0)))
                    remaining_capacity[row_key] = max(0, remaining_capacity.get(row_key, 0) - int(line.quantity or 0))

    for report in reports:
        product_name, style_name = canonical_item(session, report.company_name, report.product_name, report.style_name)
        for line in report.lines:
            if line.order_line_id:
                continue
            key = (report.company_name, product_name, style_name, line.size)
            unassigned_quantity = int(line.quantity or 0)
            allocations: list[tuple[object, dict, int]] = []
            for row in rows_by_base_key.get(key, []):
                row_key = _balance_detail_key(row)
                available = remaining_capacity.get(row_key, 0)
                if available <= 0:
                    continue
                assigned_quantity = min(available, unassigned_quantity)
                if assigned_quantity <= 0:
                    break
                allocations.append((row_key, row, assigned_quantity))
                remaining_capacity[row_key] = available - assigned_quantity
                unassigned_quantity -= assigned_quantity
                if unassigned_quantity <= 0:
                    break
            if allocations:
                detail_text = _split_shipment_detail_text(
                    report,
                    int(line.quantity or 0),
                    [(row, assigned_quantity) for _row_key, row, assigned_quantity in allocations],
                )
                for row_key, _row, _assigned_quantity in allocations:
                    details.setdefault(row_key, []).append(detail_text)
    return {key: "；".join(values) for key, values in details.items()}


def add_balance_sheet(wb: Workbook, title: str, balances: list[dict], shipment_details: dict[object, str] | None = None) -> None:
    ws = wb.create_sheet(title)
    shipment_details = shipment_details or {}
    ws.append(["公司", "产品", "款式", "订单", "尺码", "SKU", "发货明细", "下单数量", "已发数量", "未发数量", "超发数量"])
    for row in balances:
        key = _balance_detail_key(row)
        ws.append([
            row["company"], row["product"], row["style"], row.get("order_ref", ""), row["size"], row.get("sku", ""), shipment_details.get(key, ""),
            row["ordered"], row["shipped"], row["remaining"], row["over_shipped"],
        ])
    style_sheet(ws)
    return ws


def add_customer_balance_sheet(wb: Workbook, title: str, balances: list[dict], shipment_details: dict[object, str] | None = None) -> None:
    ws = wb.create_sheet(title)
    shipment_details = shipment_details or {}
    ws.append(["公司", "产品", "款式", "订单", "尺码", "SKU", "发货明细", "下单数量", "已发数量", "未发数量"])
    for row in balances:
        key = _balance_detail_key(row)
        ws.append([
            row["company"], row["product"], row["style"], row.get("order_ref", ""), row["size"], row.get("sku", ""), shipment_details.get(key, ""),
            row["ordered"], row["shipped"], row["remaining"],
        ])
    style_sheet(ws)
    return ws


def add_waybill_images_to_sheet(ws, photos) -> None:
    ws["L1"] = "快递面单"
    ws["L1"].font = Font(bold=True)
    ws["L1"].fill = PatternFill("solid", fgColor="D9EAF7")
    ws["L1"].alignment = Alignment(horizontal="center", vertical="center")
    ws.column_dimensions["L"].width = 22
    ws.column_dimensions["M"].width = 42
    row = 2
    for photo in photos:
        path = Path(photo.stored_path)
        if not path.exists():
            continue
        ws.cell(row, 12, waybill_display_name(photo))
        ws.cell(row, 12).alignment = Alignment(horizontal="center", vertical="center", wrap_text=True)
        image = ExcelImage(str(path))
        image.width = 300
        image.height = 180
        ws.add_image(image, f"M{row}")
        ws.row_dimensions[row].height = 138
        row += 9


def shipment_rows(session: Session, ship_date: str | None = None, company_name: str | None = None, user_id: int | None = None):
    query = session.query(ShipmentReport).order_by(ShipmentReport.ship_date.desc(), ShipmentReport.created_at.desc())
    if ship_date:
        query = query.filter(ShipmentReport.ship_date == ship_date)
    if company_name:
        query = query.filter(ShipmentReport.company_name == company_name)
    if user_id:
        query = query.filter(ShipmentReport.user_id == user_id)
    rows = []
    for report in query.all():
        for line in report.lines:
            rows.append([
                report.ship_date,
                report.created_at.strftime("%Y-%m-%d %H:%M"),
                report.user.display_name if report.user else "",
                report.company_name,
                report.product_name,
                report.style_name,
                line.size,
                line.quantity,
                report.status,
                report.review_reason,
                report.note,
            ])
    return rows


def add_shipments_sheet(wb: Workbook, session: Session, title: str = "发货流水", **filters) -> None:
    ws = wb.create_sheet(title)
    ws.append(["发货日期", "上报时间", "上报人", "公司", "产品", "款式", "尺码", "数量", "状态", "异常原因", "备注"])
    for row in shipment_rows(session, **filters):
        ws.append(row)
    style_sheet(ws)


def export_total_workbook(session: Session, output_path: Path) -> Path:
    wb = Workbook()
    wb.remove(wb.active)
    balances = get_order_balances(session)
    shipment_details = shipment_details_by_balance_row(session, balances)
    add_balance_sheet(wb, "订单发货明细", balances, shipment_details)
    add_shipments_sheet(wb, session)
    add_balance_sheet(wb, "未发货明细", [row for row in balances if row["remaining"] > 0], shipment_details)
    output_path.parent.mkdir(parents=True, exist_ok=True)
    wb.save(output_path)
    return output_path


def export_customer_total_workbook(session: Session, output_path: Path) -> Path:
    wb = Workbook()
    wb.remove(wb.active)
    balances = get_order_balances(session)
    shipment_details = shipment_details_by_balance_row(session, balances)
    add_customer_balance_sheet(wb, "客户发货明细", balances, shipment_details)
    add_customer_balance_sheet(wb, "未发货明细", [row for row in balances if row["remaining"] > 0], shipment_details)
    output_path.parent.mkdir(parents=True, exist_ok=True)
    wb.save(output_path)
    return output_path


def export_unshipped_workbook(session: Session, output_path: Path) -> Path:
    wb = Workbook()
    wb.remove(wb.active)
    balances = get_order_balances(session)
    add_balance_sheet(wb, "未发货明细", [row for row in balances if row["remaining"] > 0], shipment_details_by_balance_row(session, balances))
    output_path.parent.mkdir(parents=True, exist_ok=True)
    wb.save(output_path)
    return output_path


def export_company_workbook(session: Session, company_name: str, output_path: Path) -> Path:
    wb = Workbook()
    wb.remove(wb.active)
    balances = get_order_balances(session, company_name=company_name)
    ws = add_balance_sheet(wb, "订单发货明细", balances, shipment_details_by_balance_row(session, balances))
    add_waybill_images_to_sheet(ws, get_waybill_photos(session, company_name))
    add_shipments_sheet(wb, session, company_name=company_name)
    output_path.parent.mkdir(parents=True, exist_ok=True)
    wb.save(output_path)
    return output_path


def export_customer_company_workbook(session: Session, company_name: str, output_path: Path) -> Path:
    wb = Workbook()
    wb.remove(wb.active)
    balances = get_order_balances(session, company_name=company_name)
    ws = add_customer_balance_sheet(wb, "客户发货明细", balances, shipment_details_by_balance_row(session, balances))
    add_waybill_images_to_sheet(ws, get_waybill_photos(session, company_name))
    output_path.parent.mkdir(parents=True, exist_ok=True)
    wb.save(output_path)
    return output_path


def export_daily_shipments_workbook(session: Session, ship_date: str, output_path: Path) -> Path:
    wb = Workbook()
    wb.remove(wb.active)
    add_shipments_sheet(wb, session, ship_date=ship_date)
    output_path.parent.mkdir(parents=True, exist_ok=True)
    wb.save(output_path)
    return output_path


def export_employee_shipments_workbook(session: Session, user_id: int, output_path: Path) -> Path:
    wb = Workbook()
    wb.remove(wb.active)
    add_shipments_sheet(wb, session, user_id=user_id)
    output_path.parent.mkdir(parents=True, exist_ok=True)
    wb.save(output_path)
    return output_path

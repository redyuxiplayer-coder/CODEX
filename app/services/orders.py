from pathlib import Path

from openpyxl import load_workbook
from sqlalchemy import func
from sqlalchemy.orm import Session

from app.models import (
    Company,
    OrderAdjustment,
    OrderLine,
    OrderLineClose,
    ReturnRework,
    ShipmentLine,
    ShipmentReport,
    User,
)
from app.services.aliases import canonical_item
from app.services.skus import sku_lookup

APPROVED_STATUSES = {"auto_approved", "approved_after_edit"}
SIZE_SORT_ORDER = {
    "S": 0,
    "M": 1,
    "L": 2,
    "XL": 3,
    "XXL": 4,
}


def size_sort_key(size: str) -> tuple[int, str]:
    clean_size = clean_text(size).upper()
    return (SIZE_SORT_ORDER.get(clean_size, 99), clean_size)


def clean_text(value) -> str:
    if value is None:
        return ""
    return str(value).strip()


def get_or_create_company(session: Session, name: str) -> Company:
    name = clean_text(name)
    company = session.query(Company).filter_by(name=name).one_or_none()
    if company is None:
        company = Company(name=name, note="", is_active=True)
        session.add(company)
        session.flush()
    return company


def create_order_line(
    session: Session,
    company_name: str,
    product_name: str,
    style_name: str,
    size: str,
    quantity: int,
    order_date: str = "",
    delivery_date: str = "",
    note: str = "",
    batch: str = "",
) -> OrderLine:
    company = get_or_create_company(session, company_name)
    order = OrderLine(
        company_id=company.id,
        product_name=clean_text(product_name),
        style_name=clean_text(style_name),
        size=clean_text(size),
        quantity=int(quantity),
        order_date=clean_text(order_date),
        delivery_date=clean_text(delivery_date),
        note=clean_text(note),
        batch=clean_text(batch),
        is_active=True,
    )
    session.add(order)
    session.commit()
    return order


def build_structured_note(accessories: str = "", material: str = "", spec_size: str = "", note: str = "") -> str:
    parts = []
    accessories = clean_text(accessories)
    material = clean_text(material)
    spec_size = clean_text(spec_size)
    note = clean_text(note)
    if accessories:
        parts.append(f"配件：{accessories}")
    if material:
        parts.append(f"材质：{material}")
    if spec_size:
        parts.append(f"尺寸：{spec_size}")
    if note:
        parts.append(note)
    return "；".join(parts)


def clean_order_lines_from_form(sizes: list[str], quantities: list[str]) -> list[dict]:
    lines = []
    for size, quantity in zip(sizes, quantities):
        clean_size = clean_text(size)
        try:
            clean_quantity = int(quantity or 0)
        except (TypeError, ValueError):
            clean_quantity = 0
        if clean_size and clean_quantity > 0:
            lines.append({"size": clean_size, "quantity": clean_quantity})
    return lines


def find_duplicate_order_lines(
    session: Session,
    company_name: str,
    product_name: str,
    style_name: str,
    order_date: str,
    lines: list[dict],
) -> list[OrderLine]:
    company_name = clean_text(company_name)
    sizes = [line["size"] for line in lines]
    if not sizes:
        return []
    query = (
        session.query(OrderLine)
        .join(Company, Company.id == OrderLine.company_id)
        .filter(
            Company.name == company_name,
            OrderLine.product_name == clean_text(product_name),
            OrderLine.style_name == clean_text(style_name),
            OrderLine.order_date == clean_text(order_date),
            OrderLine.size.in_(sizes),
            OrderLine.is_active.is_(True),
        )
    )
    return query.order_by(OrderLine.size).all()


def create_order_lines_batch(
    session: Session,
    company_name: str,
    product_name: str,
    style_name: str,
    lines: list[dict],
    order_date: str = "",
    delivery_date: str = "",
    note: str = "",
    batch: str = "",
) -> list[OrderLine]:
    created = []
    for line in lines:
        created.append(
            create_order_line(
                session,
                company_name,
                product_name,
                style_name,
                line["size"],
                line["quantity"],
                order_date,
                delivery_date,
                note,
                batch,
            )
        )
    return created


def update_order_line(
    session: Session,
    order_id: int,
    *,
    company_name: str | None = None,
    product_name: str | None = None,
    style_name: str | None = None,
    size: str | None = None,
    quantity: int | None = None,
    order_date: str | None = None,
    delivery_date: str | None = None,
    note: str | None = None,
) -> OrderLine:
    order = session.get(OrderLine, order_id)
    if order is None:
        raise ValueError("订单不存在")
    if company_name is not None:
        order.company = get_or_create_company(session, company_name)
    if product_name is not None:
        order.product_name = clean_text(product_name)
    if style_name is not None:
        order.style_name = clean_text(style_name)
    if size is not None:
        order.size = clean_text(size)
    if quantity is not None:
        order.quantity = int(quantity)
    if order_date is not None:
        order.order_date = clean_text(order_date)
    if delivery_date is not None:
        order.delivery_date = clean_text(delivery_date)
    if note is not None:
        order.note = clean_text(note)
    session.commit()
    session.refresh(order)
    return order


def delete_order_line(session: Session, order_id: int) -> OrderLine:
    order = session.get(OrderLine, order_id)
    if order is None:
        raise ValueError("订单不存在")
    order.is_active = False
    session.commit()
    session.refresh(order)
    return order


def get_order_choices(session: Session) -> dict:
    companies: dict[str, dict[str, dict[str, set[str]]]] = {}
    for row in get_order_balances(session):
        company = row["company"]
        product = row["product"]
        style = row["style"]
        size = row["size"]
        companies.setdefault(company, {}).setdefault(product, {}).setdefault(style, set()).add(size)
    return {
        company: {
            product: {style: sorted(sizes) for style, sizes in styles.items()}
            for product, styles in products.items()
        }
        for company, products in companies.items()
    }


def approved_shipments_subquery(session: Session):
    return (
        session.query(
            ShipmentReport.company_name.label("company"),
            ShipmentReport.product_name.label("product"),
            ShipmentReport.style_name.label("style"),
            ShipmentLine.size.label("size"),
            func.sum(ShipmentLine.quantity).label("shipped"),
        )
        .join(ShipmentLine, ShipmentLine.report_id == ShipmentReport.id)
        .filter(ShipmentReport.status.in_(APPROVED_STATUSES))
        .group_by(ShipmentReport.company_name, ShipmentReport.product_name, ShipmentReport.style_name, ShipmentLine.size)
        .subquery()
    )


def get_order_balances(session: Session, company_name: str | None = None) -> list[dict]:
    order_query = (
        session.query(
            OrderLine.id.label("id"),
            Company.name.label("company"),
            OrderLine.product_name.label("product"),
            OrderLine.style_name.label("style"),
            OrderLine.size.label("size"),
            OrderLine.quantity.label("ordered"),
            OrderLine.order_date.label("order_date"),
            OrderLine.delivery_date.label("delivery_date"),
            OrderLine.batch.label("batch"),
            OrderLine.note.label("note"),
        )
        .join(Company, Company.id == OrderLine.company_id)
        .filter(OrderLine.is_active.is_(True), Company.is_active.is_(True))
    )
    if company_name:
        order_query = order_query.filter(Company.name == company_name)
    order_rows = order_query.order_by(Company.name, OrderLine.product_name, OrderLine.style_name, OrderLine.size, OrderLine.created_at, OrderLine.id).all()

    shipped_rows = (
        session.query(
            ShipmentReport.company_name,
            ShipmentReport.product_name,
            ShipmentReport.style_name,
            ShipmentLine.size,
            ShipmentLine.order_line_id,
            func.sum(ShipmentLine.quantity),
        )
        .join(ShipmentLine, ShipmentLine.report_id == ShipmentReport.id)
        .filter(ShipmentReport.status.in_(APPROVED_STATUSES))
        .group_by(ShipmentReport.company_name, ShipmentReport.product_name, ShipmentReport.style_name, ShipmentLine.size, ShipmentLine.order_line_id)
        .all()
    )
    shipped: dict[tuple[str, str, str, str], int] = {}
    shipped_by_order_id: dict[int, int] = {}
    for company, product, style, size, order_line_id, qty in shipped_rows:
        if order_line_id:
            shipped_by_order_id[int(order_line_id)] = shipped_by_order_id.get(int(order_line_id), 0) + int(qty or 0)
            continue
        canonical_product, canonical_style = canonical_item(session, company, product, style)
        key = (company, canonical_product, canonical_style, size)
        shipped[key] = shipped.get(key, 0) + int(qty or 0)

    returned_by_order_id = {
        int(order_id): int(qty or 0)
        for order_id, qty in session.query(ReturnRework.order_line_id, func.sum(ReturnRework.quantity))
        .group_by(ReturnRework.order_line_id)
        .all()
    }
    adjusted_by_order_id = {
        int(order_id): int(qty or 0)
        for order_id, qty in session.query(OrderAdjustment.order_line_id, func.sum(OrderAdjustment.quantity))
        .group_by(OrderAdjustment.order_line_id)
        .all()
    }
    scrapped_by_order_id = {
        int(order_id): int(qty or 0)
        for order_id, qty in session.query(ReturnRework.order_line_id, func.sum(ReturnRework.quantity))
        .filter(ReturnRework.status == "scrapped")
        .group_by(ReturnRework.order_line_id)
        .all()
    }
    closed_by_order_id = {
        int(order_id): int(qty or 0)
        for order_id, qty in session.query(OrderLineClose.order_line_id, func.sum(OrderLineClose.quantity))
        .group_by(OrderLineClose.order_line_id)
        .all()
    }
    skus = sku_lookup(session, company_name)

    raw_counts: dict[tuple[str, str, str, str, str, str, str], int] = {}
    raw_seen: dict[tuple[str, str, str, str, str, str, str], int] = {}
    prepared_rows = []
    for row in order_rows:
        canonical_product, canonical_style = canonical_item(session, row.company, row.product, row.style)
        raw_key = (
            row.company,
            canonical_product,
            canonical_style,
            row.product,
            row.style,
            row.size,
            row.batch or "",
        )
        raw_counts[raw_key] = raw_counts.get(raw_key, 0) + 1
        prepared_rows.append((row, canonical_product, canonical_style, raw_key))

    merged: dict[tuple[str, str, str, str, str], dict] = {}
    order_sort: dict[tuple[str, str, str, str, str], tuple[str, int]] = {}
    for row, canonical_product, canonical_style, raw_key in prepared_rows:
        raw_seen[raw_key] = raw_seen.get(raw_key, 0) + 1
        order_ref = row.batch or row.order_date or ""
        if not order_ref and raw_counts.get(raw_key, 0) > 1:
            order_ref = f"第{raw_seen[raw_key]}单"
        key = (row.company, canonical_product, canonical_style, row.size, order_ref)
        base_key = (row.company, canonical_product, canonical_style, row.size)
        ordered = int(row.ordered or 0)
        current = merged.setdefault(
            key,
            {
                "company": row.company,
                "order_id": int(row.id or 0),
                "order_ids": [int(row.id or 0)],
                "product": canonical_product,
                "style": canonical_style,
                "order_ref": order_ref,
                "order_date": row.order_date or "",
                "delivery_date": row.delivery_date or "",
                "size": row.size,
                "ordered": 0,
                "shipped": 0,
                "remaining": 0,
                "over_shipped": 0,
                "returned": 0,
                "adjusted": 0,
                "closed": 0,
                "note": row.note or "",
                "sku": skus.get(base_key, ""),
            },
        )
        order_sort.setdefault(key, (row.order_date or row.batch or "", int(row.id or 0)))
        current["ordered"] += ordered
        row_id = int(row.id or 0)
        if row_id and row_id not in current["order_ids"]:
            current["order_ids"].append(row_id)
        if row.note and row.note not in current["note"]:
            current["note"] = "；".join([text for text in [current["note"], row.note] if text])

    by_base_key: dict[tuple[str, str, str, str], list[tuple[tuple[str, str, str, str, str], dict]]] = {}
    for key, row in merged.items():
        base_key = (row["company"], row["product"], row["style"], row["size"])
        by_base_key.setdefault(base_key, []).append((key, row))

    balances = []
    for base_key, rows in by_base_key.items():
        remaining_shipped = shipped.get(base_key, 0)
        sorted_rows = sorted(rows, key=lambda item: order_sort.get(item[0], ("", 0)))
        for index, (_key, row) in enumerate(sorted_rows):
            ordered = int(row["ordered"])
            direct_shipped = sum(shipped_by_order_id.get(order_id, 0) for order_id in row.get("order_ids", []))
            open_quantity = max(0, ordered - direct_shipped)
            allocated_shipped = min(open_quantity, remaining_shipped)
            remaining_shipped -= allocated_shipped
            shipped_qty = direct_shipped + allocated_shipped
            if index == len(sorted_rows) - 1 and remaining_shipped > 0:
                shipped_qty += remaining_shipped
                remaining_shipped = 0
            returned = sum(returned_by_order_id.get(order_id, 0) for order_id in row.get("order_ids", []))
            adjusted = sum(adjusted_by_order_id.get(order_id, 0) for order_id in row.get("order_ids", [])) + sum(
                scrapped_by_order_id.get(order_id, 0) for order_id in row.get("order_ids", [])
            )
            closed = sum(closed_by_order_id.get(order_id, 0) for order_id in row.get("order_ids", []))
            remaining = ordered - shipped_qty + returned - adjusted - closed
            row["shipped"] = shipped_qty
            row["returned"] = returned
            row["adjusted"] = adjusted
            row["closed"] = closed
            row["remaining"] = remaining
            row["over_shipped"] = max(0, -remaining)
            balances.append(row)
    return sorted(
        balances,
        key=lambda x: (
            x["company"],
            x["product"],
            x["style"],
            x.get("order_date") or x.get("order_ref") or "",
            size_sort_key(x["size"]),
        ),
    )


def find_header_row(ws):
    for row in ws.iter_rows():
        values = [clean_text(cell.value) for cell in row]
        if "公司" in values and "款式" in values and "尺码" in values and "下单数量" in values:
            return row[0].row, values
    return None, []


def preview_excel_orders(path: Path) -> dict:
    wb = load_workbook(path, data_only=True)
    ws = wb["订单发货明细"] if "订单发货明细" in wb.sheetnames else wb.active
    header_row, headers = find_header_row(ws)
    if not header_row:
        return {"valid_rows": [], "errors": [{"row": 0, "message": "没有找到订单表头"}]}
    idx = {name: headers.index(name) + 1 for name in headers if name}
    valid_rows = []
    errors = []
    for row_no in range(header_row + 1, ws.max_row + 1):
        company = clean_text(ws.cell(row_no, idx.get("公司", 0)).value)
        product = clean_text(ws.cell(row_no, idx.get("产品", 0)).value)
        style = clean_text(ws.cell(row_no, idx.get("款式", 0)).value)
        size = clean_text(ws.cell(row_no, idx.get("尺码", 0)).value)
        quantity_raw = ws.cell(row_no, idx.get("下单数量", 0)).value
        if not any([company, product, style, size, quantity_raw]):
            continue
        try:
            quantity = int(float(quantity_raw))
        except (TypeError, ValueError):
            errors.append({"row": row_no, "message": "下单数量不是数字"})
            continue
        if not company or not style or not size or quantity <= 0:
            errors.append({"row": row_no, "message": "公司、款式、尺码不能为空，数量必须大于0"})
            continue
        valid_rows.append(
            {
                "row": row_no,
                "company": company,
                "product": product or style,
                "style": style,
                "size": size,
                "quantity": quantity,
                "shipped": _safe_int(ws.cell(row_no, idx.get("已发合计", 0)).value) if "已发合计" in idx else 0,
                "shipping_record": clean_text(ws.cell(row_no, idx.get("发货记录", 0)).value) if "发货记录" in idx else "",
                "note": clean_text(ws.cell(row_no, idx.get("备注", 0)).value) if "备注" in idx else "",
                "batch": clean_text(ws.cell(row_no, idx.get("订单批次", 0)).value) if "订单批次" in idx else "",
            }
        )
    return {"valid_rows": valid_rows, "errors": errors}


def _safe_int(value) -> int:
    try:
        return int(float(value or 0))
    except (TypeError, ValueError):
        return 0


def import_excel_orders(session: Session, path: Path) -> dict:
    preview = preview_excel_orders(path)
    imported = 0
    history_user = session.query(User).filter_by(role="admin").first() or session.query(User).first()
    for row in preview["valid_rows"]:
        order = create_order_line(
            session,
            row["company"],
            row["product"],
            row["style"],
            row["size"],
            row["quantity"],
            note=row["note"],
            batch=row["batch"],
        )
        if history_user and row.get("shipped", 0) > 0:
            report = ShipmentReport(
                user_id=history_user.id,
                ship_date="历史导入",
                company_name=row["company"],
                product_name=row["product"],
                style_name=row["style"],
                note=row.get("shipping_record") or "从旧Excel导入已发数量",
                status="auto_approved",
                review_reason="",
            )
            session.add(report)
            session.flush()
            session.add(ShipmentLine(report_id=report.id, order_line_id=order.id, size=row["size"], quantity=int(row["shipped"])))
            session.commit()
        imported += 1
    return {"imported": imported, "errors": preview["errors"]}

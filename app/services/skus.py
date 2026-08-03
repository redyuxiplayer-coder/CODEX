from datetime import datetime
from pathlib import Path

from openpyxl import load_workbook
from sqlalchemy.orm import Session

from app.models import SkuMapping


def clean_text(value) -> str:
    if value is None:
        return ""
    return str(value).strip()


def upsert_sku_mapping(
    session: Session,
    company_name: str,
    product_name: str,
    style_name: str,
    size: str,
    sku: str,
) -> SkuMapping:
    company_name = clean_text(company_name)
    product_name = clean_text(product_name)
    style_name = clean_text(style_name)
    size = clean_text(size)
    sku = clean_text(sku)
    mapping = (
        session.query(SkuMapping)
        .filter_by(company_name=company_name, product_name=product_name, style_name=style_name, size=size)
        .one_or_none()
    )
    if mapping is None:
        mapping = SkuMapping(
            company_name=company_name,
            product_name=product_name,
            style_name=style_name,
            size=size,
            sku=sku,
        )
        session.add(mapping)
    else:
        mapping.sku = sku
        mapping.updated_at = datetime.now()
    session.commit()
    session.refresh(mapping)
    return mapping


def sku_lookup(session: Session, company_name: str | None = None) -> dict[tuple[str, str, str, str], str]:
    query = session.query(SkuMapping)
    if company_name:
        query = query.filter(SkuMapping.company_name == company_name)
    return {
        (row.company_name, row.product_name, row.style_name, row.size): row.sku
        for row in query.all()
        if row.sku
    }


def import_sku_mappings_from_excel(session: Session, path: Path) -> dict[str, int]:
    wb = load_workbook(path, data_only=True)
    ws = wb["源兴发SKU对应表"] if "源兴发SKU对应表" in wb.sheetnames else wb.active
    header_row = None
    headers: list[str] = []
    for row in ws.iter_rows():
        values = [clean_text(cell.value) for cell in row]
        if {"公司", "产品", "款式", "尺码", "SKU"}.issubset(set(values)):
            header_row = row[0].row
            headers = values
            break
    if header_row is None:
        raise ValueError("没有找到 SKU 表头：公司、产品、款式、尺码、SKU")

    idx = {name: headers.index(name) + 1 for name in headers if name}
    imported = 0
    skipped = 0
    for row_no in range(header_row + 1, ws.max_row + 1):
        company = clean_text(ws.cell(row_no, idx["公司"]).value)
        product = clean_text(ws.cell(row_no, idx["产品"]).value)
        style = clean_text(ws.cell(row_no, idx["款式"]).value)
        size = clean_text(ws.cell(row_no, idx["尺码"]).value)
        sku = clean_text(ws.cell(row_no, idx["SKU"]).value)
        if not any([company, product, style, size, sku]):
            continue
        if not all([company, product, style, size, sku]):
            skipped += 1
            continue
        upsert_sku_mapping(session, company, product, style, size, sku)
        imported += 1
    return {"imported": imported, "skipped": skipped}
